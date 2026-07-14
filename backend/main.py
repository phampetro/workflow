"""
PyFlow Studio — Backend (Flask)
Python 3.14 compatible, không cần pydantic/fastapi

Chạy: .venv\\Scripts\\python main.py
API docs: Xem README.md
"""
import os
import json
import uuid
import sqlite3
import logging
import urllib.request
import urllib.error
import urllib.parse
import smtplib
from email.message import EmailMessage
import asyncio
import subprocess
import sys
import threading
import unicodedata
import re
from datetime import datetime
from pathlib import Path
import platform

from flask import Flask, request, jsonify, g, send_file
from flask_cors import CORS
from apscheduler.schedulers.background import BackgroundScheduler
from croniter import croniter

# ── Config ──────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "pyflow.db"

DATA_DIR.mkdir(exist_ok=True)

# ── Logging ─────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("pyflow")

# ── Flask App ───────────────────────────────────────────────
app = Flask(__name__)
CORS(app, origins=["http://localhost:5173", "http://127.0.0.1:5173"])

# ── APScheduler ─────────────────────────────────────────────
scheduler = BackgroundScheduler()

# WebSocket log clients (run_id -> list of queues)
_log_clients: dict = {}
_active_runs: dict = {}  # run_id -> stop_event
_active_procs: dict = {}  # run_id -> subprocess.Popen (for force-kill)
_workflow_run_ids: dict = {}  # workflow_id -> set of run_ids
_run_logs_cache: dict = {} # run_id -> list of historical logs


# ── Database ─────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db:
        db.close()


def init_db():
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS user (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT,
            is_active INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS project (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            color TEXT DEFAULT '#6c63ff',
            created_at TEXT,
            updated_at TEXT,
            venv_ready INTEGER DEFAULT 0,
            venv_path TEXT,
            user_id TEXT,
            FOREIGN KEY (user_id) REFERENCES user(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS workflow (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            project_id TEXT NOT NULL,
            created_at TEXT,
            updated_at TEXT,
            graph_json TEXT,
            FOREIGN KEY (project_id) REFERENCES project(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS schedule (
            id TEXT PRIMARY KEY,
            workflow_id TEXT NOT NULL,
            cron_expr TEXT NOT NULL,
            label TEXT,
            enabled INTEGER DEFAULT 1,
            created_at TEXT,
            next_run_at TEXT,
            last_run_at TEXT,
            FOREIGN KEY (workflow_id) REFERENCES workflow(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS workflow_run (
            id TEXT PRIMARY KEY,
            workflow_id TEXT NOT NULL,
            project_id TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            started_at TEXT,
            finished_at TEXT,
            duration_ms INTEGER,
            triggered_by TEXT DEFAULT 'manual',
            error_message TEXT,
            FOREIGN KEY (workflow_id) REFERENCES workflow(id) ON DELETE CASCADE
        );
        """)

        # Safe migrations — bỏ qua nếu cột đã tồn tại
        migrations = [
            "ALTER TABLE project ADD COLUMN user_id TEXT",
            "ALTER TABLE user ADD COLUMN is_active INTEGER DEFAULT 0",
            "ALTER TABLE project ADD COLUMN icon TEXT DEFAULT 'Box'",
            "ALTER TABLE workflow ADD COLUMN color TEXT DEFAULT '#6c63ff'",
            "ALTER TABLE project ADD COLUMN sort_order INTEGER DEFAULT 0",
            "ALTER TABLE workflow ADD COLUMN sort_order INTEGER DEFAULT 0",
        ]
        for sql in migrations:
            try:
                conn.execute(sql)
            except Exception:
                pass  # Column already exists
        conn.commit()
    logger.info("✅ Database đã sẵn sàng")


def cleanup_stale_runs():
    """Khi backend khởi động lại, đánh dấu tất cả run đang 'running' là 'error'.
    Những run này bị interrupt do server restart nên không thể hoàn thành."""
    with sqlite3.connect(str(DB_PATH)) as conn:
        stale = conn.execute(
            "SELECT id FROM workflow_run WHERE status='running'"
        ).fetchall()
        if stale:
            now = now_iso()
            conn.execute(
                """UPDATE workflow_run 
                   SET status='error', finished_at=?, error_message='Workflow bị gián đoạn do server restart' 
                   WHERE status='running'""",
                (now,)
            )
            conn.commit()
            logger.warning(f"⚠️  Đã cleanup {len(stale)} run bị treo (stale) từ lần chạy trước")
        else:
            logger.info("✅ Không có stale runs")




def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def now_iso() -> str:
    """Return current local time in ISO 8601 format."""
    return datetime.now().astimezone().isoformat()

# ══════════════════════════════════════════════════════════════
#  VEnv Manager & Directory Structure
# ══════════════════════════════════════════════════════════════

def slugify(text: str) -> str:
    if not text:
        return "untitled"
    text = unicodedata.normalize('NFKD', str(text)).encode('ascii', 'ignore').decode('utf-8')
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    text = re.sub(r'[\s_-]+', '_', text)
    return text

def get_project_dir(project_id: str) -> Path:
    with sqlite3.connect(str(DB_PATH)) as conn:
        row = conn.execute("SELECT name FROM project WHERE id=?", (project_id,)).fetchone()
        name = row[0] if row else "unknown"
    return DATA_DIR / f"pj_{slugify(name)}"

def get_venv_dir(project_id: str) -> Path:
    return get_project_dir(project_id) / "venv"

def get_python_path(project_id: str) -> str:
    venv = get_venv_dir(project_id)
    if sys.platform == "win32":
        return str(venv / "Scripts" / "python.exe")
    return str(venv / "bin" / "python")


def get_pip_path(project_id: str) -> str:
    venv = get_venv_dir(project_id)
    if sys.platform == "win32":
        return str(venv / "Scripts" / "pip.exe")
    return str(venv / "bin" / "pip")


def venv_exists(project_id: str) -> bool:
    return os.path.isfile(get_python_path(project_id))


def create_venv_sync(project_id: str) -> dict:
    venv_path = get_venv_dir(project_id)
    venv_path.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [sys.executable, "-m", "venv", str(venv_path)],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0:
        raise RuntimeError(f"Không tạo được venv: {result.stderr}")
    return {"path": str(venv_path), "python": get_python_path(project_id)}


def install_pkg_sync(project_id: str, package: str) -> dict:
    if not venv_exists(project_id):
        create_venv_sync(project_id)
    pip = get_pip_path(project_id)
    result = subprocess.run([pip, "install", package, "--quiet"], capture_output=True, text=True, timeout=300)
    if result.returncode != 0:
        raise RuntimeError(f"Cài thất bại: {result.stderr[:500]}")
    return {"package": package, "status": "installed"}


def uninstall_pkg_sync(project_id: str, package: str) -> dict:
    pip = get_pip_path(project_id)
    subprocess.run([pip, "uninstall", package, "-y"], capture_output=True, text=True, timeout=60)
    return {"package": package, "status": "uninstalled"}

def ensure_packages(project_id: str, packages: list, log_fn=None, bid=None, label=""):
    if not packages: return
    if not venv_exists(project_id):
        create_venv_sync(project_id)
    
    pip = get_pip_path(project_id)
    result = subprocess.run([pip, "freeze"], capture_output=True, text=True)
    installed = result.stdout.lower()
    
    missing = []
    for pkg in packages:
        pkg_name = pkg.lower().split('==')[0].split('>')[0].split('<')[0]
        if f"{pkg_name}==" not in installed and f"{pkg_name} @" not in installed:
            missing.append(pkg)
            
    if missing:
        if log_fn:
            log_fn(bid, "info", f"📦 [{label}] Đang tải & tự động cài đặt thư viện: {', '.join(missing)}...")
        
        res = subprocess.run([pip, "install"] + missing + ["--quiet"], capture_output=True, text=True, timeout=300)
        if res.returncode != 0:
            if log_fn:
                log_fn(bid, "error", f"❌ Cài đặt thất bại: {res.stderr[:200]}")
            raise RuntimeError(f"Không thể cài đặt {missing}")

def list_pkgs_sync(project_id: str) -> list:
    if not venv_exists(project_id):
        return []
    pip = get_pip_path(project_id)
    result = subprocess.run([pip, "list", "--format=json"], capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        return []
    try:
        return [{"name": p["name"], "version": p["version"]} for p in json.loads(result.stdout)]
    except Exception:
        return []


def delete_venv_sync(project_id: str):
    import shutil
    venv_path = get_venv_dir(project_id)
    if venv_path.exists():
        shutil.rmtree(venv_path, ignore_errors=True)


# ══════════════════════════════════════════════════════════════
#  Execution Engine
# ══════════════════════════════════════════════════════════════

RUNNER_TEMPLATE = '''
import sys, json, traceback, os

workflow_id = {workflow_id!r}
block_id = {block_id!r}
OUTPUT_DIR = {output_dir!r}
INPUT_DIR = {input_dir!r}

input_data = None
try:
    _raw = sys.stdin.read().strip()
    if _raw:
        input_data = json.loads(_raw)
except Exception:
    pass

output_data = None

try:
{user_code}
except Exception as _e:
    print("[ERROR] " + traceback.format_exc(), file=sys.stderr)
    sys.exit(1)

try:
    _out = json.dumps(output_data, ensure_ascii=False, default=str)
except Exception:
    _out = json.dumps(str(output_data))
print("__OUTPUT__:" + _out)
'''


def indent_code(code: str, spaces: int = 4) -> str:
    return "\n".join(" " * spaces + line for line in code.splitlines())


def topological_sort(nodes: list, edges: list) -> list:
    adj = {n["id"]: [] for n in nodes}
    in_deg = {n["id"]: 0 for n in nodes}
    for edge in edges:
        src, tgt = edge.get("source"), edge.get("target")
        if src in adj and tgt in adj:
            adj[src].append(tgt)
            in_deg[tgt] += 1
    queue = [nid for nid, d in in_deg.items() if d == 0]
    ordered = []
    while queue:
        nid = queue.pop(0)
        ordered.append(nid)
        for nxt in adj.get(nid, []):
            in_deg[nxt] -= 1
            if in_deg[nxt] == 0:
                queue.append(nxt)
    node_map = {n["id"]: n for n in nodes}
    return [node_map[nid] for nid in ordered if nid in node_map]


def get_workflow_dir(project_id: str, workflow_id: str) -> Path:
    with sqlite3.connect(str(DB_PATH)) as conn:
        row = conn.execute("SELECT name FROM workflow WHERE id=?", (workflow_id,)).fetchone()
        name = row[0] if row else "unknown"
    return get_project_dir(project_id) / f"wf_{slugify(name)}"

def run_python_block_sync(project_id, block_id, workflow_id, code, input_data, timeout=60, label=None, log_fn=None, input_dir=None, stop_event=None):
    """Chạy 1 block Python synchronously, có thể bị ngắt bởi stop_event"""
    if not venv_exists(project_id):
        create_venv_sync(project_id)

    python_exe = get_python_path(project_id)
    
    wf_dir = get_workflow_dir(project_id, workflow_id)
    wf_dir.mkdir(parents=True, exist_ok=True)
    
    block_dir = wf_dir / slugify(label or block_id)
    block_dir.mkdir(parents=True, exist_ok=True)
    
    output_dir = wf_dir / "output"
    output_dir.mkdir(exist_ok=True)
    
    block_path = block_dir / "main.py"
    wrapped = RUNNER_TEMPLATE.format(
        workflow_id=workflow_id,
        block_id=block_id,
        output_dir=str(output_dir).replace('\\', '/'),
        input_dir=str(input_dir).replace('\\', '/') if input_dir else '',
        user_code=indent_code(code),
    )
    block_path.write_text(wrapped, encoding="utf-8")

    input_json = json.dumps(input_data, ensure_ascii=False, default=str)

    if log_fn:
        log_fn(block_id, "info", f"▶  Chạy block [{label or block_id}]")

    start = datetime.now()
    run_id_for_proc = None
    # Tìm run_id tương ứng dựa vào stop_event
    for rid, ev in list(_active_runs.items()):
        if ev is stop_event:
            run_id_for_proc = rid
            break

    try:
        proc = subprocess.Popen(
            [python_exe, "-u", str(block_path)],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        # Đăng ký proc để có thể force-kill
        if run_id_for_proc:
            _active_procs[run_id_for_proc] = proc

        try:
            proc.stdin.write(input_json.encode("utf-8"))
            proc.stdin.close()
        except Exception:
            pass

        # Chờ proc và check stop_event định kỳ
        deadline = datetime.now().timestamp() + timeout
        while True:
            # Chờ tối đa 0.5s mỗi lần
            try:
                proc.wait(timeout=0.5)
                break  # proc kết thúc
            except subprocess.TimeoutExpired:
                pass

            # Kiểm tra stop_event
            if stop_event and stop_event.is_set():
                import signal
                try:
                    proc.kill()
                except Exception:
                    pass
                duration = int((datetime.now() - start).total_seconds() * 1000)
                if log_fn:
                    log_fn(block_id, "warning", f"⏹ Block đã bị dừng ({duration}ms)")
                if run_id_for_proc:
                    _active_procs.pop(run_id_for_proc, None)
                return False, None, "stopped", duration

            # Kiểm tra timeout
            if datetime.now().timestamp() > deadline:
                try:
                    proc.kill()
                except Exception:
                    pass
                duration = int((datetime.now() - start).total_seconds() * 1000)
                msg = f"⏰ Timeout sau {timeout}s"
                if log_fn:
                    log_fn(block_id, "error", msg)
                if run_id_for_proc:
                    _active_procs.pop(run_id_for_proc, None)
                return False, None, msg, duration

        if run_id_for_proc:
            _active_procs.pop(run_id_for_proc, None)

        duration = int((datetime.now() - start).total_seconds() * 1000)
        output_data = None

        stdout_bytes, stderr_bytes = proc.stdout.read(), proc.stderr.read()
        stdout_str = stdout_bytes.decode("utf-8", errors="replace")
        stderr_str = stderr_bytes.decode("utf-8", errors="replace")

        for line in stdout_str.splitlines():
            if line.startswith("__OUTPUT__:"):
                try:
                    output_data = json.loads(line[len("__OUTPUT__:"):])
                except Exception:
                    output_data = line[len("__OUTPUT__:"):]
            elif line.strip() and log_fn:
                log_fn(block_id, "info", f"   {line}")

        if stderr_str and stderr_str.strip():
            for line in stderr_str.splitlines():
                if line.strip() and log_fn:
                    log_fn(block_id, "error", f"   ⚠ {line}")

        if proc.returncode != 0:
            err = stderr_str.strip() or "Unknown error"
            if log_fn:
                log_fn(block_id, "error", f"✗ Block thất bại ({duration}ms)")
            return False, None, err, duration

        if log_fn:
            log_fn(block_id, "success", f"✓ Block hoàn thành ({duration}ms)")
        return True, output_data, None, duration

    except Exception as e:
        if run_id_for_proc:
            _active_procs.pop(run_id_for_proc, None)
        duration = int((datetime.now() - start).total_seconds() * 1000)
        if log_fn:
            log_fn(block_id, "error", f"✗ Lỗi: {e}")
        return False, None, str(e), duration


def execute_workflow_thread(run_id, project_id, workflow_id, workflow_name, graph_json, log_fn, stop_event):
    """Chạy toàn bộ workflow trong thread riêng"""
    import time
    time.sleep(0.5) # Đợi client SSE kết nối trước khi chạy nhanh
    start = datetime.now()
    try:
        graph = json.loads(graph_json)
    except Exception as e:
        _finish_run(run_id, "error", start, error=str(e))
        return
        
    proj_dir = get_project_dir(project_id)
    wf_dir = proj_dir / f"wf_{slugify(workflow_name)}"
    input_dir = wf_dir / "input"

    try:
        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])
        
        nodes_dict = {n["id"]: n for n in nodes}
        edges_from = {}
        for e in edges:
            edges_from.setdefault(e["source"], []).append(e)

        start_nodes = [n for n in nodes if n.get("data", {}).get("type") == "start"]
        if not start_nodes:
            if log_fn:
                log_fn("system", "error", "❌ Không tìm thấy khối Bắt đầu!")
            _finish_run(run_id, "error", start, error="Missing Start block")
            return

        from collections import deque
        queue = deque()
        # Enqueue: (node_id, input_data)
        queue.append((start_nodes[0]["id"], None))

        if log_fn:
            log_fn("system", "info", f"🚀 Bắt đầu workflow (Dynamic Routing)")

        final_status = "success"
        visited = set()

        while queue:
            if stop_event and stop_event.is_set():
                log_fn("system", "warning", "⏹ Đã dừng bởi người dùng")
                final_status = "stopped"
                break

            node_id, current_input = queue.popleft()
            if node_id in visited:
                continue
            visited.add(node_id)

            node = nodes_dict.get(node_id)
            if not node:
                continue

            bdata = node.get("data", {})
            
            # Đọc input.json một lần duy nhất nếu chưa đọc
            if "workflow_env" not in locals():
                workflow_env = {}
                try:
                    input_file = input_dir / "input.json"
                    if input_file.exists():
                        with open(input_file, "r", encoding="utf-8") as f:
                            workflow_env = json.load(f)
                except Exception:
                    pass
                    
                def interpolate(val):
                    if not isinstance(val, str):
                        return val
                    if val in workflow_env:
                        return str(workflow_env[val])
                    for k, v in workflow_env.items():
                        val = val.replace("{{" + k + "}}", str(v))
                    return val
            
            # Nội suy các biến môi trường
            bdata_interpolated = {}
            for k, v in bdata.items():
                if isinstance(v, str):
                    bdata_interpolated[k] = interpolate(v)
                else:
                    bdata_interpolated[k] = v
            bdata = bdata_interpolated

            btype = bdata.get("type", "python")
            bid = node["id"]
            label = bdata.get("label", bid)

            # Execution flags for branching
            continue_branch = True
            cond_branch_taken = None # 'true' or 'false'

            if btype == "start":
                if log_fn:
                    log_fn(bid, "info", f"▶  [Start] {label}")
            elif btype == "end":
                if log_fn:
                    log_fn(bid, "info", f"🏁 [End] {label}")
            elif btype == "delay":
                delay_sec = float(bdata.get("delaySeconds", 3))
                if log_fn:
                    log_fn(bid, "info", f"⏳ [Delay] {label} - Đang chờ {delay_sec} giây...")
                import time
                time.sleep(delay_sec)
                if log_fn:
                    log_fn(bid, "success", f"✅ Đã chờ xong {delay_sec} giây.")
            elif btype == "telegram":
                bot_token = bdata.get("telegramBotToken", "").strip()
                chat_id = bdata.get("telegramChatId", "").strip()
                message_template = bdata.get("telegramMessage", "")
                parse_mode = bdata.get("telegramParseMode", "")
                
                if not bot_token or not chat_id:
                    if log_fn:
                        log_fn(bid, "error", f"❌ [Telegram] {label} - Thiếu Bot Token hoặc Chat ID")
                    final_status = "error"
                    break
                    
                text = message_template
                try:
                    text = text.replace("{input_data}", str(current_input))
                    if isinstance(current_input, dict):
                        for k, v in current_input.items():
                            text = text.replace("{" + str(k) + "}", str(v))
                except Exception:
                    pass
                    
                if log_fn:
                    log_fn(bid, "info", f"✉️ [Telegram] {label} - Đang gửi tin nhắn tới {chat_id}...")
                
                url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                payload = {"chat_id": chat_id, "text": text}
                if parse_mode:
                    payload["parse_mode"] = parse_mode
                    
                req_data = json.dumps(payload).encode('utf-8')
                req = urllib.request.Request(url, data=req_data, headers={'Content-Type': 'application/json'})
                
                try:
                    with urllib.request.urlopen(req, timeout=10) as response:
                        res = json.loads(response.read().decode('utf-8'))
                        if res.get("ok"):
                            if log_fn:
                                log_fn(bid, "success", f"✅ [Telegram] {label} - Đã gửi thành công!")
                        else:
                            if log_fn:
                                log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi API: {res.get('description')}")
                            final_status = "error"
                            break
                except Exception as e:
                    if log_fn:
                        log_fn(bid, "error", f"❌ [Telegram] {label} - Gửi thất bại: {str(e)}")
                    final_status = "error"
                    break
            elif btype == "email":
                mail_host = bdata.get("mailHost", "").strip()
                mail_port = int(bdata.get("mailPort", 465) or 465)
                mail_user = bdata.get("mailUser", "").strip()
                mail_pass = bdata.get("mailPass", "").replace(" ", "")
                mail_to = bdata.get("mailTo", "").strip()
                mail_cc = bdata.get("mailCc", "").strip()
                mail_subject = bdata.get("mailSubject", "").strip()
                mail_body = bdata.get("mailBody", "")
                mail_attachments = bdata.get("mailAttachments", [])

                def tpl(txt):
                    if not txt: return ""
                    t = txt.replace("{input_data}", str(current_input))
                    if isinstance(current_input, dict):
                        for k, v in current_input.items():
                            t = t.replace("{" + str(k) + "}", str(v))
                    return t

                final_to = tpl(mail_to)
                final_cc = tpl(mail_cc)
                final_subject = tpl(mail_subject)
                final_body = tpl(mail_body)

                if log_fn:
                    log_fn(bid, "info", f"📧 [Email] {label} - Đang gửi thư tới {final_to}...")

                msg = EmailMessage()
                msg['Subject'] = final_subject
                msg['From'] = mail_user
                msg['To'] = final_to
                if final_cc:
                    msg['Cc'] = final_cc
                
                # Check if body contains HTML tags
                if "<" in final_body and ">" in final_body:
                    msg.set_content(final_body, subtype='html')
                else:
                    msg.set_content(final_body)

                # Attachments
                for att in mail_attachments:
                    att_name = tpl(att)
                    if not att_name: continue
                    # try INPUT_DIR then OUTPUT_DIR
                    att_path = input_dir / att_name
                    if not att_path.exists():
                        att_path = wf_dir / "output" / att_name
                    
                    if att_path.exists() and att_path.is_file():
                        import mimetypes
                        ctype, encoding = mimetypes.guess_type(str(att_path))
                        if ctype is None or encoding is not None:
                            ctype = 'application/octet-stream'
                        maintype, subtype = ctype.split('/', 1)
                        with open(att_path, 'rb') as fp:
                            msg.add_attachment(fp.read(), maintype=maintype, subtype=subtype, filename=att_name)
                    else:
                        if log_fn:
                            log_fn(bid, "warning", f"⚠ [Email] Không tìm thấy file đính kèm: {att_name}")

                try:
                    # Decide SSL or TLS based on port
                    if mail_port == 465:
                        smtp = smtplib.SMTP_SSL(mail_host, mail_port, timeout=15)
                    else:
                        smtp = smtplib.SMTP(mail_host, mail_port, timeout=15)
                        smtp.starttls()
                    
                    smtp.login(mail_user, mail_pass)
                    smtp.send_message(msg)
                    smtp.quit()

                    if log_fn:
                        log_fn(bid, "success", f"✅ [Email] {label} - Đã gửi thư thành công!")
                except Exception as e:
                    if log_fn:
                        log_fn(bid, "error", f"❌ [Email] {label} - Lỗi gửi thư: {str(e)}")
                    final_status = "error"
                    break
            elif btype == "database":
                db_type = bdata.get("dbType", "postgresql")
                db_host = bdata.get("dbHost", "")
                db_port = bdata.get("dbPort", "")
                db_user = bdata.get("dbUser", "")
                db_pass = bdata.get("dbPassword", "")
                db_name = bdata.get("dbName", "")
                
                db_user_enc = urllib.parse.quote_plus(db_user) if db_user else ""
                db_pass_enc = urllib.parse.quote_plus(db_pass) if db_pass else ""

                conn_str = ""
                if db_type == "postgresql":
                    conn_str = f"postgresql://{db_user_enc}:{db_pass_enc}@{db_host}:{db_port}/{db_name}"
                elif db_type == "mysql":
                    conn_str = f"mysql+pymysql://{db_user_enc}:{db_pass_enc}@{db_host}:{db_port}/{db_name}"
                elif db_type == "sqlite":
                    conn_str = f"sqlite:///{db_name}"
                elif db_type == "sqlserver":
                    conn_str = f"mssql+pyodbc://{db_user_enc}:{db_pass_enc}@{db_host}:{db_port}/{db_name}?driver=ODBC+Driver+17+for+SQL+Server"
                    
                current_input = {
                    "db_type": db_type,
                    "host": db_host,
                    "port": db_port,
                    "user": db_user,
                    "password": db_pass,
                    "db_name": db_name,
                    "connection_string": conn_str
                }
                
                packages_to_install = []
                if db_type == "postgresql": packages_to_install.append("psycopg2-binary")
                elif db_type == "mysql": packages_to_install.extend(["pymysql", "cryptography"])
                elif db_type == "sqlserver": packages_to_install.append("pyodbc")
                
                ensure_packages(project_id, packages_to_install, log_fn, bid, label)
                
                if log_fn:
                    log_fn(bid, "success", f"✅ [Database] {label} - Đã tạo cấu hình kết nối {db_type.upper()}")
            elif btype == "browser":
                steps = bdata.get("steps", [])
                headless = not bdata.get("debugMode", False)
                if not steps:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có bước nào, bỏ qua")
                    continue_branch = False
                else:
                    if log_fn:
                        log_fn(bid, "info", f"🌐 Đang chạy Browser: {label}...")
                    
                    import asyncio
                    from services.browser_executor import run_browser_block
                    
                    # Wrap sync log_fn to async
                    async def async_log_cb(b_id, lvl, msg):
                        if log_fn:
                            log_fn(b_id, lvl, msg)
                            
                    start_b = datetime.now()
                    try:
                        output_dir = wf_dir / "output"
                        output_dir.mkdir(exist_ok=True)
                        b_result = asyncio.run(run_browser_block(
                            block_id=bid,
                            workflow_id=workflow_id,
                            steps=steps,
                            input_data=current_input,
                            headless=headless,
                            log_callback=async_log_cb,
                            output_dir=str(output_dir).replace('\\', '/'),
                        ))
                        if not b_result.get("success"):
                            if log_fn:
                                log_fn("system", "error", f"❌ Workflow thất bại (Browser lỗi)")
                            _finish_run(run_id, "error", start, error=b_result.get("error"))
                            return
                        else:
                            if b_result.get("output_data") is not None:
                                current_input = b_result["output_data"]
                    except Exception as e:
                        if log_fn:
                            log_fn("system", "error", f"❌ Lỗi ngoại lệ Browser: {e}")
                        _finish_run(run_id, "error", start, error=str(e))
                        return
            elif btype == "python":
                code = bdata.get("code", "").strip()
                if not code:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có code")
                    continue_branch = False
                else:
                    if log_fn:
                        log_fn(bid, "info", f"⚡ Đang chạy: {label}...")
                    success, output, error, duration = run_python_block_sync(
                        project_id, bid, workflow_id, code, current_input, 
                        timeout=1800, label=label, log_fn=log_fn, input_dir=str(input_dir),
                        stop_event=stop_event
                    )
                    if not success:
                        if error == "stopped":
                            final_status = "stopped"
                            break
                        _finish_run(run_id, "error", start, error=error)
                        if log_fn:
                            log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                        return
                    current_input = output
            elif btype == "sql_to_excel":
                sql_query = bdata.get("sqlQuery", "").strip()
                excel_filename = bdata.get("excelFileName", "export.xlsx").strip()
                if not excel_filename:
                    excel_filename = "export.xlsx"
                
                if not sql_query:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có câu lệnh SQL")
                    continue_branch = False
                else:
                    packages_to_install = ["pandas", "sqlalchemy", "openpyxl"]
                    if isinstance(current_input, dict):
                        db_type = current_input.get("db_type")
                        if db_type == "postgresql": packages_to_install.append("psycopg2-binary")
                        elif db_type == "mysql": packages_to_install.extend(["pymysql", "cryptography"])
                        elif db_type == "sqlserver": packages_to_install.append("pyodbc")
                    
                    ensure_packages(project_id, packages_to_install, log_fn, bid, label)
                    
                    if log_fn:
                        log_fn(bid, "info", f"⚡ Đang chạy SQL to Excel: {label}...")
                    
                    code = f'''
import pandas as pd
import sqlalchemy
import os

conn_str = input_data.get("connection_string")
if not conn_str:
    raise ValueError("Không tìm thấy connection_string từ khối trước. Hãy đảm bảo nối khối này vào sau khối Database.")

print("Đang kết nối CSDL và thực thi câu lệnh SQL...")
engine = sqlalchemy.create_engine(conn_str)
df = pd.read_sql("""{sql_query}""", engine)

file_name = {excel_filename!r}
out_path = os.path.join(OUTPUT_DIR, file_name)

print(f"Đã tải {{len(df)}} dòng dữ liệu! Đang lưu vào {{out_path}}...")
df.to_excel(out_path, index=False)

output_data = {{"status": "success", "file_path": out_path}}
'''
                    success, output, error, duration = run_python_block_sync(
                        project_id, bid, workflow_id, code, current_input, 
                        timeout=1800, label=label, log_fn=log_fn, input_dir=str(input_dir)
                    )
                    if not success:
                        _finish_run(run_id, "error", start, error=error)
                        if log_fn:
                            log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                        return
                    current_input = output
            elif btype == "merge_excel":
                header_rows = int(bdata.get("headerRows", 3))
                excel_filename = bdata.get("excelFileName", "merged.xlsx").strip()
                merge_all_input = bdata.get("mergeAllInput", True)
                selected_files = bdata.get("selectedFiles", [])
                if not excel_filename:
                    excel_filename = "merged.xlsx"
                
                # Nếu bật chọn tất cả Input: tự động quét toàn bộ file trong INPUT_DIR
                if merge_all_input:
                    selected_files = None  # Sẽ tự quét trong code Python
                elif not selected_files:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Merge Excel: Bạn chưa chọn file nào để gộp!")
                    _finish_run(run_id, "error", start, error="No files selected")
                    return
                
                packages_to_install = ["pandas", "openpyxl"]
                ensure_packages(project_id, packages_to_install, log_fn, bid, label)
                
                if log_fn:
                    log_fn(bid, "info", f"⚡ Đang chạy Merge Excel: {label}...")
                
                if merge_all_input:
                    file_list_code = """
# T\u1ef1 đ\u1ed9ng qu\u00e9t t\u1ea5t c\u1ea3 file .xlsx/.csv trong INPUT_DIR
file_list = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx') or f.endswith('.csv')])
print(f"T\u1ef1 đ\u1ed9ng ph\u00e1t hi\u1ec7n {len(file_list)} file trong th\u01b0 m\u1ee5c Input.")
"""
                else:
                    _fl = selected_files
                    file_list_code = f"""file_list = {_fl!r}\nprint(f\"\u0110ang gh\u00e9p {{len(file_list)}} file \u0111\u00e3 ch\u1ecdn.\")"""

                code = f'''
import os
import pandas as pd
import openpyxl

{file_list_code}

if not file_list:
    raise ValueError(f"Không tìm thấy file .xlsx nào trong Dữ liệu Workflow (INPUT_DIR)")

print(f"Bắt đầu ghép. Giữ nguyên {{ {header_rows} }} dòng tiêu đề từ file đầu tiên...")

file1_path = os.path.join(INPUT_DIR, file_list[0])
if not os.path.exists(file1_path):
    raise FileNotFoundError(f"Không tìm thấy file Gốc: {{file_list[0]}}")

# 1. Dùng openpyxl mở File 1 để giữ trọn vẹn Format tiêu đề
print(f"Đang xử lý định dạng tiêu đề từ File 1: {{file_list[0]}}")
wb = openpyxl.load_workbook(file1_path)
ws = wb.active

# Xóa toàn bộ dữ liệu bên dưới dòng tiêu đề (Để lấy format chuẩn cho phần data sau này)
if ws.max_row > {header_rows}:
    ws.delete_rows({header_rows} + 1, ws.max_row)

# 2. Dùng pandas đọc phần Data của TẤT CẢ các file (Bao gồm cả file 1)
dfs = []
for filename in file_list:
    file_path = os.path.join(INPUT_DIR, filename)
    print(f"   + Đang đọc dữ liệu: {{filename}}")
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Không tìm thấy file: {{filename}}")
        
    # Bỏ qua header_rows, chỉ lấy Data thô
    df = pd.read_excel(file_path, header=None, skiprows={header_rows})
    dfs.append(df)

print("Đang gộp và đắp dữ liệu vào bảng...")
merged_df = pd.concat(dfs, ignore_index=True)

# 3. Ghi dữ liệu vào Workbook đã có sẵn tiêu đề đẹp
# Chuyển DataFrame thành list các dòng để append vào openpyxl siêu nhanh
for row in merged_df.itertuples(index=False, name=None):
    ws.append(row)

out_file = {excel_filename!r}
out_path = os.path.join(OUTPUT_DIR, out_file)

print(f"Đang lưu file kết quả (Gồm {header_rows} dòng tiêu đề + {{len(merged_df)}} dòng dữ liệu) vào {{out_path}}...")
try:
    wb.save(out_path)
except PermissionError:
    raise PermissionError(f"Không thể lưu đè file {{out_file}} do file này đang được mở trong Excel. Vui lòng đóng file Excel và chạy lại.")
output_data = {{"status": "success", "file_path": out_path}}
'''
                success, output, error, duration = run_python_block_sync(
                    project_id, bid, workflow_id, code, current_input, 
                    timeout=1800, label=label, log_fn=log_fn, input_dir=str(input_dir)
                )
                if not success:
                    _finish_run(run_id, "error", start, error=error)
                    if log_fn:
                        log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                    return
                current_input = output
            elif btype == "pivot_excel":
                excel_filename = bdata.get("excelFileName", "pivot_result.xlsx").strip()
                selected_files = bdata.get("pivotInputFiles", [])
                pivot_index = bdata.get("pivotIndex", "")
                pivot_columns = bdata.get("pivotColumns", "")
                pivot_values = bdata.get("pivotValues", "")
                pivot_agg = bdata.get("pivotAgg", "sum")
                pivot_fillna = bdata.get("pivotFillNa", True)
                pivot_grand_total = bdata.get("pivotGrandTotal", True)
                pivot_header_row = int(bdata.get("pivotHeaderRow", 1))
                pivot_enable_sort = bdata.get("pivotEnableSort", False)
                pivot_sort_column = bdata.get("pivotSortColumn", "").strip()
                pivot_sort_order = bdata.get("pivotSortOrder", "asc")
                pivot_sort_custom = bdata.get("pivotSortCustom", [])
                
                if not excel_filename:
                    excel_filename = "pivot_result.xlsx"
                
                if not selected_files:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Pivot Excel: Bạn chưa chọn file nào để tổng hợp!")
                    _finish_run(run_id, "error", start, error="No files selected")
                    return
                
                packages_to_install = ["pandas", "openpyxl"]
                ensure_packages(project_id, packages_to_install, log_fn, bid, label)
                
                if log_fn:
                    log_fn(bid, "info", f"📊 Đang chạy Pivot Excel: {label}...")
                
                code = f'''
import os
import pandas as pd

print(f"Đang tìm kiếm {{len({selected_files!r})}} file Excel trong thư mục: {{INPUT_DIR}}...")
file_list = {selected_files!r}

if not file_list:
    raise ValueError(f"Không tìm thấy file nào để xử lý.")

dfs = []
for filename in file_list:
    file_path = os.path.join(INPUT_DIR, filename)
    if not os.path.exists(file_path):
        file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Không tìm thấy file: {{filename}}")
    
    print(f"Đang đọc dữ liệu: {{filename}} (header dòng {{ {pivot_header_row} }})")
    pd_header_idx = max(0, {pivot_header_row} - 1)
    if filename.endswith(".csv"):
        df = pd.read_csv(file_path, header=pd_header_idx)
    else:
        df = pd.read_excel(file_path, header=pd_header_idx)
    dfs.append(df)

df = pd.concat(dfs, ignore_index=True)

def col2num(col_str):
    num = 0
    for c in str(col_str).upper():
        if 'A' <= c <= 'Z':
            num = num * 26 + (ord(c) - ord('A') + 1)
        else:
            return -1
    return num - 1

# Lọc và dọn dẹp biến đầu vào
def parse_cols(col_data):
    if not col_data: return []
    if isinstance(col_data, str):
        cols = [c.strip() for c in col_data.split(',')]
    elif isinstance(col_data, list):
        cols = [str(c).strip() for c in col_data]
    else:
        return []
        
    parsed = []
    for c in cols:
        if not c: continue
        if c in df.columns:
            parsed.append(c)
        else:
            idx = col2num(c)
            if 0 <= idx < len(df.columns):
                parsed.append(df.columns[idx])
            else:
                print(f"Cảnh báo: Không tìm thấy cột '{{c}}'")
    return parsed

idx_cols = parse_cols({pivot_index!r})
col_cols = parse_cols({pivot_columns!r})
val_cols = parse_cols({pivot_values!r})

if not val_cols:
    print("Cảnh báo: Không có cột Giá trị (Values) nào hợp lệ. Sẽ dùng toàn bộ cột số.")
    val_cols = df.select_dtypes(include='number').columns.tolist()
if not idx_cols and not col_cols:
    raise ValueError("Phải nhập ít nhất 1 trường Dòng (Rows) hoặc Cột (Columns).")

sort_col_parsed = parse_cols({pivot_sort_column!r})
sort_col = sort_col_parsed[0] if sort_col_parsed else None

if {pivot_enable_sort} and sort_col and sort_col in df.columns:
    print(f"Đang phân loại và sắp xếp cột '{{sort_col}}' theo chiều: {{ {pivot_sort_order!r} }}")
    if {pivot_sort_order!r} == 'custom':
        custom_order = {pivot_sort_custom!r}
        # Xử lý Pandas4Warning: bổ sung các giá trị còn thiếu vào cuối danh sách categories
        unique_vals = df[sort_col].dropna().unique().tolist()
        # Ép kiểu custom_order về cùng kiểu với dữ liệu để so sánh chính xác nếu cần
        try:
            custom_order = pd.Series(custom_order).astype(df[sort_col].dtype).tolist()
        except:
            pass
        missing_cats = [x for x in unique_vals if x not in custom_order]
        custom_order.extend(missing_cats)
        df[sort_col] = pd.Categorical(df[sort_col], categories=custom_order, ordered=True)
    elif {pivot_sort_order!r} == 'asc':
        cats = sorted(df[sort_col].dropna().unique().tolist(), key=str)
        df[sort_col] = pd.Categorical(df[sort_col], categories=cats, ordered=True)
    elif {pivot_sort_order!r} == 'desc':
        cats = sorted(df[sort_col].dropna().unique().tolist(), key=str, reverse=True)
        df[sort_col] = pd.Categorical(df[sort_col], categories=cats, ordered=True)

print(f"Đang Pivot với Dòng={{idx_cols}}, Cột={{col_cols}}, Giá trị={{val_cols}}, Hàm={pivot_agg!r}")

pivot_df = pd.pivot_table(
    df, 
    values=val_cols, 
    index=idx_cols if idx_cols else None, 
    columns=col_cols if col_cols else None, 
    aggfunc={pivot_agg!r}, 
    margins={pivot_grand_total},
    margins_name="Tổng Cộng"
)

if {pivot_fillna}:
    pivot_df = pivot_df.fillna(0)

# Làm phẳng cột đa cấp (nếu có)
if isinstance(pivot_df.columns, pd.MultiIndex):
    new_cols = []
    for col in pivot_df.columns.values:
        valid_parts = [str(c) for c in col if str(c) != '']
        if not valid_parts:
            new_cols.append('')
        elif len(valid_parts) > 1 and valid_parts[0] in val_cols:
            new_cols.append('_'.join(valid_parts[1:]))
        else:
            new_cols.append('_'.join(valid_parts))
    pivot_df.columns = new_cols

out_file = {excel_filename!r}
out_path = os.path.join(OUTPUT_DIR, out_file)

# Reset index để lưu ra file excel đẹp hơn
pivot_df.reset_index(inplace=True)

print(f"Đang lưu file Pivot kết quả vào {{out_path}}...")
try:
    pivot_df.to_excel(out_path, index=False)
except PermissionError:
    raise PermissionError(f"Không thể lưu đè file {{out_file}} do file này đang được mở trong Excel.")

output_data = {{"status": "success", "file_path": out_path}}
'''
                success, output, error, duration = run_python_block_sync(
                    project_id, bid, workflow_id, code, current_input, 
                    timeout=1800, label=label, log_fn=log_fn, input_dir=str(input_dir)
                )
                if not success:
                    _finish_run(run_id, "error", start, error=error)
                    if log_fn:
                        log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                    return
                current_input = output
            elif btype == "condition":
                cond_var = bdata.get("condVariable", "").strip()
                cond_op = bdata.get("condOperator", "==")
                cond_val = bdata.get("condValue", "").strip()
                
                if log_fn:
                    log_fn(bid, "info", f"🤔 Đang kiểm tra: {cond_var} {cond_op} {cond_val}")
                
                try:
                    # Lấy giá trị biến từ current_input (là dictionary)
                    actual_val = None
                    if isinstance(current_input, dict):
                        actual_val = current_input.get(cond_var)
                    
                    # So sánh động
                    result = False
                    
                    # Convert cond_val to same type if possible for comparison
                    cmp_val = cond_val
                    if isinstance(actual_val, int):
                        try: cmp_val = int(cond_val)
                        except: pass
                    elif isinstance(actual_val, float):
                        try: cmp_val = float(cond_val)
                        except: pass
                    elif isinstance(actual_val, bool):
                        cmp_val = cond_val.lower() == 'true'
                        
                    if cond_op == "==":
                        result = (actual_val == cmp_val) or (str(actual_val) == str(cmp_val))
                    elif cond_op == "!=":
                        result = (actual_val != cmp_val) and (str(actual_val) != str(cmp_val))
                    elif cond_op == ">":
                        result = float(actual_val) > float(cmp_val)
                    elif cond_op == "<":
                        result = float(actual_val) < float(cmp_val)
                    elif cond_op == ">=":
                        result = float(actual_val) >= float(cmp_val)
                    elif cond_op == "<=":
                        result = float(actual_val) <= float(cmp_val)
                    elif cond_op == "contains":
                        result = str(cmp_val) in str(actual_val)
                        
                    cond_branch_taken = "true" if result else "false"
                    if log_fn:
                        log_fn(bid, "success", f"✅ Kết quả: '{cond_var}' có giá trị là '{actual_val}' ➜ {actual_val} {cond_op} {cmp_val} là {result}")
                except Exception as e:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Lỗi so sánh: {e}")
                    final_status = "error"
                    break

            if continue_branch and final_status != "error":
                out_edges = edges_from.get(node_id, [])
                for e in out_edges:
                    target_id = e["target"]
                    source_handle = e.get("sourceHandle")
                    
                    # Xử lý rẽ nhánh điều kiện
                    if btype == "condition":
                        if source_handle == cond_branch_taken:
                            queue.append((target_id, current_input))
                    else:
                        queue.append((target_id, current_input))

        total_ms = int((datetime.now() - start).total_seconds() * 1000)
        _finish_run(run_id, final_status, start)
        if log_fn:
            if final_status == "success":
                log_fn("system", "success", f"✅ Workflow hoàn thành trong {total_ms}ms")
            elif final_status == "stopped":
                log_fn("system", "warning", f"⏹ Đã dừng sau {total_ms}ms")
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        _finish_run(run_id, "error", start, error=str(e))
        if log_fn:
            log_fn("system", "error", f"❌ Lỗi hệ thống khi chạy workflow: {str(e)}")
        print(f"CRITICAL ERROR IN WORKFLOW THREAD: {err_msg}")


def _finish_run(run_id, status, start, error=None):
    finished = datetime.now()
    duration = int((finished - start).total_seconds() * 1000)
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.execute(
            "UPDATE workflow_run SET status=?, finished_at=?, duration_ms=?, error_message=? WHERE id=?",
            (status, finished.isoformat(), duration, error, run_id)
        )
    _active_runs.pop(run_id, None)
    _active_procs.pop(run_id, None)
    # Xóa run_id khỏi workflow mapping
    for wf_id, run_set in list(_workflow_run_ids.items()):
        run_set.discard(run_id)
        if not run_set:
            _workflow_run_ids.pop(wf_id, None)
    
    # Gửi None để đóng luồng stream
    queues = _log_clients.get(run_id, [])
    for q in queues:
        try:
            q.put_nowait(None)
        except Exception:
            pass


def make_log_fn(run_id):
    """Tạo log function để ghi vào DB và broadcast tới SSE clients"""
    def log_fn(block_id, level, message):
        payload = json.dumps({
            "run_id": run_id,
            "block_id": block_id,
            "level": level,
            "message": message,
            "time": datetime.now().strftime("%H:%M:%S"),
        })
        _run_logs_cache.setdefault(run_id, []).append(payload)
        # Push tới tất cả clients đang lắng nghe run này
        queues = _log_clients.get(run_id, [])
        for q in list(queues):
            try:
                q.put_nowait(payload)
            except Exception:
                pass
    return log_fn




def get_next_run(cron_expr: str):
    try:
        from apscheduler.triggers.cron import CronTrigger
        trigger = CronTrigger(**_cron_kwargs(cron_expr))
        # Use system local timezone for computing next run
        next_run = trigger.get_next_fire_time(None, datetime.now(datetime.now().astimezone().tzinfo))
        return next_run.isoformat() if next_run else None
    except Exception as e:
        logger.error(f"Error getting next run: {e}")
        return None


def _scheduler_trigger(workflow_id, project_id, schedule_id):
    """Gọi khi scheduler trigger — chạy workflow"""
    logger.info(f"⏰ Scheduler: workflow={workflow_id}")
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        wf = conn.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
        if not wf or not wf["graph_json"]:
            return

        sched = conn.execute("SELECT cron_expr FROM schedule WHERE id=?", (schedule_id,)).fetchone()
        
        run_id = str(uuid.uuid4())
        now = now_iso()
        conn.execute(
            "INSERT INTO workflow_run (id, workflow_id, project_id, status, started_at, triggered_by) VALUES (?,?,?,?,?,?)",
            (run_id, workflow_id, project_id, "running", now, f"schedule:{schedule_id}")
        )
        next_run = get_next_run(sched["cron_expr"]) if sched else None
        conn.execute("UPDATE schedule SET last_run_at=?, next_run_at=? WHERE id=?", (now, next_run, schedule_id))

    stop_event = threading.Event()
    _active_runs[run_id] = stop_event
    log_fn = make_log_fn(run_id)

    t = threading.Thread(
        target=execute_workflow_thread,
        args=(run_id, project_id, workflow_id, wf["name"], wf["graph_json"], log_fn, stop_event),
        daemon=True
    )
    t.start()


def load_schedules_from_db():
    """Reload schedules vào APScheduler khi khởi động — CHỈ load của user đang is_active=1"""
    with sqlite3.connect(str(DB_PATH)) as conn:
        conn.row_factory = sqlite3.Row
        # Tìm user đang active
        active_user = conn.execute("SELECT id, name FROM user WHERE is_active=1").fetchone()
        if not active_user:
            logger.info("ℹ️  Không có user active → không load schedules")
            return
        logger.info(f"👤 Active user: {active_user['name']} — đang load schedules...")
        rows = conn.execute("""
            SELECT s.*, w.project_id FROM schedule s
            JOIN workflow w ON s.workflow_id = w.id
            JOIN project p ON w.project_id = p.id
            WHERE s.enabled = 1 AND p.user_id = ?
        """, (active_user["id"],)).fetchall()
        for row in rows:
            try:
                scheduler.add_job(
                    _scheduler_trigger,
                    "cron",
                    id=row["id"],
                    kwargs={"workflow_id": row["workflow_id"], "project_id": row["project_id"], "schedule_id": row["id"]},
                    **_cron_kwargs(row["cron_expr"]),
                    replace_existing=True,
                )
                logger.info(f"  ↻ Loaded schedule: {row['label']} ({row['cron_expr']})")
            except Exception as e:
                logger.warning(f"  ⚠ Lỗi load schedule {row['id']}: {e}")
        logger.info(f"✅ Loaded {len(rows)} schedules cho user '{active_user['name']}'")


def _cron_kwargs(expr: str) -> dict:
    """Parse cron expression (or JSON) thành APScheduler kwargs"""
    if expr.startswith("{"):
        config = json.loads(expr)
        kwargs = {
            "hour": config.get("hour", "*"),
            "minute": config.get("minute", "*"),
        }
        
        if config.get("schedule_type") == "month":
            kwargs["day"] = str(config.get("day_of_month", "1"))
        else:
            kwargs["day_of_week"] = ",".join(config.get("days", [])) or "*"

        if config.get("start_date"):
            kwargs["start_date"] = config.get("start_date")
        if config.get("end_date"):
            kwargs["end_date"] = config.get("end_date")
        return kwargs
    else:
        parts = expr.split()
        if len(parts) != 5:
            raise ValueError(f"Cron không hợp lệ: {expr}")
        keys = ["minute", "hour", "day", "month", "day_of_week"]
        return dict(zip(keys, parts))


# ══════════════════════════════════════════════════════════════
#  API Routes — Projects
# ══════════════════════════════════════════════════════════════

@app.route("/api/projects", methods=["GET"])
def list_projects():
    db = get_db()
    user_id = request.headers.get('X-User-Id')
    if user_id:
        rows = db.execute("""
            SELECT p.*, 
                   (SELECT COUNT(*) FROM workflow WHERE project_id = p.id) as workflow_count,
                   (SELECT COUNT(*) FROM workflow_run WHERE project_id = p.id AND status = 'running') as running_count
            FROM project p WHERE p.user_id = ? ORDER BY p.sort_order ASC, p.created_at DESC
        """, (user_id,)).fetchall()
    else:
        rows = db.execute("""
            SELECT p.*, 
                   (SELECT COUNT(*) FROM workflow WHERE project_id = p.id) as workflow_count,
                   (SELECT COUNT(*) FROM workflow_run WHERE project_id = p.id AND status = 'running') as running_count
            FROM project p ORDER BY p.sort_order ASC, p.created_at DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/projects", methods=["POST"])
def create_project():
    body = request.get_json(force=True)
    user_id = request.headers.get('X-User-Id')
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Tên project không được để trống"}), 400
    pid = str(uuid.uuid4())
    now = now_iso()
    db = get_db()
    # Kiểm tra tên trùng
    dup = db.execute(
        "SELECT id FROM project WHERE name=? AND user_id IS ?",
        (name, user_id)
    ).fetchone()
    if dup:
        return jsonify({"error": f"Tên project '{name}' đã tồn tại"}), 409
    db.execute(
        "INSERT INTO project (id, name, description, color, icon, created_at, updated_at, user_id) VALUES (?,?,?,?,?,?,?,?)",
        (pid, name, body.get("description"), body.get("color", "#6c63ff"), body.get("icon", "Box"), now, now, user_id)
    )
    db.commit()
    # Tạo venv trong background
    threading.Thread(target=_create_venv_bg, args=(pid,), daemon=True).start()

    row = db.execute("SELECT * FROM project WHERE id=?", (pid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/projects/<project_id>", methods=["PUT"])
def update_project(project_id):
    body = request.get_json(force=True)
    db = get_db()
    project = db.execute("SELECT * FROM project WHERE id=?", (project_id,)).fetchone()
    if not project:
        return jsonify({"error": "Project not found"}), 404

    old_name = project["name"]
    new_name = (body.get("name") or old_name).strip()
    user_id = project["user_id"]

    # Kiểm tra tên trùng (bỏ qua chính nó)
    if new_name != old_name:
        dup = db.execute(
            "SELECT id FROM project WHERE name=? AND user_id IS ? AND id != ?",
            (new_name, user_id, project_id)
        ).fetchone()
        if dup:
            return jsonify({"error": f"Tên project '{new_name}' đã tồn tại"}), 409

    db.execute(
        "UPDATE project SET name=?, description=?, color=?, icon=?, updated_at=? WHERE id=?",
        (
            new_name,
            body.get("description", project["description"]),
            body.get("color", project["color"]),
            body.get("icon", dict(project).get("icon", "Box")),
            now_iso(),
            project_id
        )
    )
    db.commit()
    row = db.execute("SELECT * FROM project WHERE id=?", (project_id,)).fetchone()

    if new_name != old_name:
        old_dir = DATA_DIR / f"pj_{slugify(old_name)}"
        new_dir = DATA_DIR / f"pj_{slugify(new_name)}"
        if old_dir.exists() and not new_dir.exists():
            old_dir.rename(new_dir)

    return jsonify(dict(row))


def _create_venv_bg(project_id):
    try:
        result = create_venv_sync(project_id)
        with sqlite3.connect(str(DB_PATH)) as conn:
            conn.execute(
                "UPDATE project SET venv_ready=1, venv_path=?, updated_at=? WHERE id=?",
                (result["path"], now_iso(), project_id)
            )
        logger.info(f"✅ Venv tạo xong cho project {project_id}")
    except Exception as e:
        logger.error(f"❌ Lỗi tạo venv cho {project_id}: {e}")


@app.route("/api/projects/<project_id>", methods=["GET"])
def get_project(project_id):
    db = get_db()
    row = db.execute("SELECT * FROM project WHERE id=?", (project_id,)).fetchone()
    if not row:
        return jsonify({"error": "Project không tồn tại"}), 404
    return jsonify(dict(row))





@app.route("/api/projects/<project_id>", methods=["DELETE"])
def delete_project(project_id):
    db = get_db()
    import shutil
    proj_dir = get_project_dir(project_id)
    if proj_dir.exists():
        shutil.rmtree(proj_dir, ignore_errors=True)
        
    db.execute("DELETE FROM project WHERE id=?", (project_id,))
    db.commit()
    return "", 204


@app.route("/api/dashboard/stats", methods=["GET"])
def get_dashboard_stats():
    db = get_db()
    user_id = request.headers.get('X-User-Id')

    if user_id:
        total_projects = db.execute(
            "SELECT COUNT(*) as c FROM project WHERE user_id=?", (user_id,)
        ).fetchone()["c"]
        total_workflows = db.execute(
            "SELECT COUNT(*) as c FROM workflow WHERE project_id IN (SELECT id FROM project WHERE user_id=?)",
            (user_id,)
        ).fetchone()["c"]
        schedule_filter = "AND SUBSTR(triggered_by, 10) IN (SELECT s.id FROM schedule s JOIN workflow w ON s.workflow_id=w.id JOIN project p ON w.project_id=p.id WHERE p.user_id=?)"
        schedule_params = (f"{datetime.now().strftime('%Y-%m-%d')}%", user_id)
    else:
        total_projects = db.execute("SELECT COUNT(*) as c FROM project").fetchone()["c"]
        total_workflows = db.execute("SELECT COUNT(*) as c FROM workflow").fetchone()["c"]
        schedule_filter = "AND SUBSTR(triggered_by, 10) IN (SELECT id FROM schedule)"
        schedule_params = (f"{datetime.now().strftime('%Y-%m-%d')}%",)

    # Số tiến trình đang chạy
    running_count = len(_active_runs)

    # Lịch đã thực thi trong hôm nay
    today_executed = db.execute(
        f"SELECT COUNT(*) as c FROM workflow_run WHERE started_at LIKE ? AND triggered_by LIKE 'schedule:%' {schedule_filter}",
        schedule_params
    ).fetchone()["c"]

    # Lịch sẽ thực thi hôm nay
    today_local = datetime.now().strftime("%Y-%m-%d")
    today_remaining = 0
    for job in scheduler.get_jobs():
        if job.next_run_time:
            if job.next_run_time.strftime("%Y-%m-%d") == today_local:
                today_remaining += 1

    return jsonify({
        "status": "ok",
        "data": {
            "total_projects": total_projects,
            "total_workflows": total_workflows,
            "running": running_count,
            "today_executed": today_executed,
            "today_remaining": today_remaining,
            "today_total": today_executed + today_remaining
        }
    })



@app.route("/api/projects/<project_id>/packages", methods=["GET"])
def get_packages(project_id):
    return jsonify(list_pkgs_sync(project_id))


@app.route("/api/projects/<project_id>/packages/install", methods=["POST"])
def install_package(project_id):
    body = request.get_json(force=True)
    pkg = (body.get("package") or "").strip()
    if not pkg:
        return jsonify({"error": "Thiếu tên package"}), 400
    try:
        result = install_pkg_sync(project_id, pkg)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/projects/<project_id>/packages/uninstall", methods=["POST"])
def uninstall_package(project_id):
    body = request.get_json(force=True)
    pkg = (body.get("package") or "").strip()
    if not pkg:
        return jsonify({"error": "Thiếu tên package"}), 400
    result = uninstall_pkg_sync(project_id, pkg)
    return jsonify(result)


@app.route("/api/projects/<project_id>/venv/init", methods=["POST"])
def init_venv(project_id):
    try:
        result = create_venv_sync(project_id)
        db = get_db()
        db.execute("UPDATE project SET venv_ready=1, venv_path=? WHERE id=?", (result["path"], project_id))
        db.commit()
        return jsonify({"status": "ok", "path": result["path"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  API Routes — Workflows
# ══════════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/workflows", methods=["GET"])
def list_workflows(project_id):
    db = get_db()
    rows = db.execute('''
        SELECT w.*, 
               (SELECT COUNT(1) FROM workflow_run r WHERE r.workflow_id = w.id AND r.status = 'running') as running_count
        FROM workflow w 
        WHERE w.project_id=? 
        ORDER BY w.sort_order ASC, w.created_at DESC
    ''', (project_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/projects/reorder", methods=["PUT"])
def reorder_projects():
    """Cập nhật sort_order cho danh sách projects."""
    items = request.get_json(force=True)  # [{id, sort_order}, ...]
    db = get_db()
    for item in items:
        db.execute("UPDATE project SET sort_order=? WHERE id=?", (item["sort_order"], item["id"]))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>/workflows/reorder", methods=["PUT"])
def reorder_workflows(project_id):
    """Cập nhật sort_order cho danh sách workflows trong project."""
    items = request.get_json(force=True)  # [{id, sort_order}, ...]
    db = get_db()
    for item in items:
        db.execute("UPDATE workflow SET sort_order=? WHERE id=? AND project_id=?",
                   (item["sort_order"], item["id"], project_id))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>/workflows", methods=["POST"])
def create_workflow(project_id):
    body = request.get_json(force=True)
    db = get_db()
    proj = db.execute("SELECT id FROM project WHERE id=?", (project_id,)).fetchone()
    if not proj:
        return jsonify({"error": "Project không tồn tại"}), 404

    wid = str(uuid.uuid4())
    now = now_iso()
    wf_name = (body.get("name") or "").strip()
    if not wf_name:
        return jsonify({"error": "Tên workflow không được để trống"}), 400
    wf_color = body.get("color", "#6c63ff")
    # Kiểm tra tên trùng trong cùng project
    dup = db.execute(
        "SELECT id FROM workflow WHERE name=? AND project_id=?",
        (wf_name, project_id)
    ).fetchone()
    if dup:
        return jsonify({"error": f"Tên workflow '{wf_name}' đã tồn tại trong project này"}), 409
    db.execute(
        "INSERT INTO workflow (id, name, description, color, project_id, created_at, updated_at) VALUES (?,?,?,?,?,?,?)",
        (wid, wf_name, body.get("description"), wf_color, project_id, now, now)
    )
    db.commit()
    
    # Tạo sẵn thư mục workflow
    proj_dir = get_project_dir(project_id)
    wf_dir = proj_dir / f"wf_{slugify(wf_name)}"
    wf_dir.mkdir(parents=True, exist_ok=True)
    
    row = db.execute("SELECT * FROM workflow WHERE id=?", (wid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/workflows/<workflow_id>", methods=["GET"])
def get_workflow(workflow_id):
    db = get_db()
    row = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not row:
        return jsonify({"error": "Workflow không tồn tại"}), 404
    return jsonify(dict(row))


@app.route("/api/workflows/<workflow_id>/duplicate", methods=["POST"])
def duplicate_workflow(workflow_id):
    db = get_db()
    row = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not row:
        return jsonify({"error": "Workflow không tồn tại"}), 404

    new_id = str(uuid.uuid4())
    new_name = row["name"] + " (Copy)"
    now = datetime.utcnow().isoformat()
    
    db.execute(
        """
        INSERT INTO workflow (id, name, description, project_id, created_at, updated_at, graph_json, color, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (new_id, new_name, row["description"], row["project_id"], now, now, row["graph_json"], row["color"], row["sort_order"])
    )
    db.commit()
    
    new_row = db.execute("SELECT * FROM workflow WHERE id=?", (new_id,)).fetchone()
    return jsonify(dict(new_row)), 201


@app.route("/api/workflows/<workflow_id>", methods=["PUT"])
def update_workflow(workflow_id):
    body = request.get_json(force=True)
    db = get_db()
    
    old_row = db.execute("SELECT name, project_id FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    old_name = old_row["name"] if old_row else None
    proj_id = old_row["project_id"] if old_row else None

    # Kiểm tra tên trùng nếu đang đổi tên
    new_name = (body.get("name") or "").strip() if "name" in body else None
    if new_name and new_name != old_name and proj_id:
        dup = db.execute(
            "SELECT id FROM workflow WHERE name=? AND project_id=? AND id!=?",
            (new_name, proj_id, workflow_id)
        ).fetchone()
        if dup:
            return jsonify({"error": f"Tên workflow '{new_name}' đã tồn tại trong project này"}), 409

    fields = {k: body[k] for k in ["name", "description", "graph_json", "color"] if k in body}
    if not fields:
        return jsonify({"error": "Không có dữ liệu"}), 400
    fields["updated_at"] = now_iso()
    sets = ", ".join(f"{k}=?" for k in fields)
    db.execute(f"UPDATE workflow SET {sets} WHERE id=?", (*fields.values(), workflow_id))
    db.commit()
    row = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    
    if proj_id:
        proj_dir = get_project_dir(proj_id)
        new_wf_name = fields.get("name", old_name)
        new_dir = proj_dir / f"wf_{slugify(new_wf_name)}"
        
        if old_name and "name" in fields and fields["name"] != old_name:
            old_dir = proj_dir / f"wf_{slugify(old_name)}"
            if old_dir.exists() and not new_dir.exists():
                old_dir.rename(new_dir)
                
        # Trích xuất và lưu block code ngay khi save
        if "graph_json" in fields and fields["graph_json"]:
            try:
                graph = json.loads(fields["graph_json"])
                new_dir.mkdir(parents=True, exist_ok=True)
                for node in graph.get("nodes", []):
                    if node.get("data", {}).get("type", "python") == "python":
                        bid = node["id"]
                        label = node["data"].get("label", bid)
                        code = node["data"].get("code", "")
                        block_dir = new_dir / slugify(label)
                        block_dir.mkdir(parents=True, exist_ok=True)
                        input_dir = new_dir / "input"
                        output_dir = new_dir / "output"
                        input_dir.mkdir(exist_ok=True)
                        output_dir.mkdir(exist_ok=True)
                        
                        block_path = block_dir / "main.py"
                        wrapped = RUNNER_TEMPLATE.format(
                            workflow_id=workflow_id,
                            block_id=bid,
                            input_dir=str(input_dir).replace('\\', '/'),
                            output_dir=str(output_dir).replace('\\', '/'),
                            user_code=indent_code(code)
                        )
                        block_path.write_text(wrapped, encoding="utf-8")
            except Exception as e:
                logger.error(f"Lỗi extract block python: {e}")

    return jsonify(dict(row))


@app.route("/api/workflows/<workflow_id>/input", methods=["GET"])
def get_workflow_input(workflow_id):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_file = wf_dir / "input" / "input.json"
    
    if not input_file.exists():
        return jsonify({"TELEGRAM_TOKEN": "", "TELEGRAM_CHAT_ID": ""})
        
    try:
        with open(input_file, "r", encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception:
        return jsonify({})

@app.route("/api/workflows/<workflow_id>/input", methods=["PUT"])
def update_workflow_input(workflow_id):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    data = request.json
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_dir = wf_dir / "input"
    input_dir.mkdir(parents=True, exist_ok=True)
    
    input_file = input_dir / "input.json"
    with open(input_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
        
    return jsonify({"success": True})

@app.route("/api/workflows/<workflow_id>/files", methods=["GET"])
def list_workflow_files(workflow_id):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_dir = wf_dir / "input"
    output_dir = wf_dir / "output"
    
    files = []
    seen = set()
    
    # Files from input
    if input_dir.exists():
        for f in input_dir.iterdir():
            if f.is_file() and f.name != "input.json":
                stat = f.stat()
                files.append({
                    "name": f.name,
                    "size": stat.st_size,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "source": "input"
                })
                seen.add(f.name)
                
    # Files from output
    if output_dir.exists():
        for f in output_dir.iterdir():
            if f.is_file() and f.name not in seen:
                stat = f.stat()
                files.append({
                    "name": f.name,
                    "size": stat.st_size,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "source": "output"
                })
    return jsonify(files)

@app.route("/api/workflows/<workflow_id>/file-columns", methods=["GET"])
def get_workflow_file_columns(workflow_id):
    filename = request.args.get("filename")
    header_row = int(request.args.get("header_row", 1))
    
    if not filename:
        return jsonify({"error": "No filename"}), 400
        
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    
    file_path = wf_dir / "input" / filename
    if not file_path.exists():
        file_path = wf_dir / "output" / filename
        
    if not file_path.exists():
        return jsonify({"error": "File not found"}), 404
        
    try:
        import pandas as pd
        header_idx = max(0, header_row - 1)
        if str(file_path).endswith(".csv"):
            df = pd.read_csv(str(file_path), header=header_idx, nrows=0)
        else:
            try:
                df = pd.read_excel(str(file_path), header=header_idx, nrows=0, engine="calamine")
            except:
                df = pd.read_excel(str(file_path), header=header_idx, nrows=0)
        return jsonify([str(c) for c in df.columns])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/workflows/<workflow_id>/file-column-values", methods=["GET"])
def get_workflow_file_column_values(workflow_id):
    filename = request.args.get("filename")
    col_name = request.args.get("col_name")
    header_row = int(request.args.get("header_row", 1))
    
    if not filename or not col_name:
        return jsonify({"error": "No filename or col_name"}), 400
        
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    
    file_path = wf_dir / "input" / filename
    if not file_path.exists():
        file_path = wf_dir / "output" / filename
        
    if not file_path.exists():
        return jsonify({"error": "File not found"}), 404
        
    try:
        import pandas as pd
        header_idx = max(0, header_row - 1)
        
        # Đọc trước header để lấy index của cột
        if str(file_path).endswith(".csv"):
            df_cols = pd.read_csv(str(file_path), header=header_idx, nrows=0)
        else:
            try:
                df_cols = pd.read_excel(str(file_path), header=header_idx, nrows=0, engine="calamine")
            except:
                df_cols = pd.read_excel(str(file_path), header=header_idx, nrows=0)
            
        str_cols = [str(c) for c in df_cols.columns]
        
        col_index = -1
        if col_name in str_cols:
            col_index = str_cols.index(col_name)
        else:
            # Hỗ trợ chữ cái cột A, B, C...
            num = 0
            for c in str(col_name).upper():
                if 'A' <= c <= 'Z':
                    num = num * 26 + (ord(c) - ord('A') + 1)
                else:
                    num = -1
                    break
            if num > 0:
                col_index = num - 1
                
        if col_index < 0 or col_index >= len(str_cols):
            return jsonify({"error": "Column not found"}), 404
        
        # Đọc chính xác cột theo index
        if str(file_path).endswith(".csv"):
            df = pd.read_csv(str(file_path), header=header_idx, usecols=[col_index])
        else:
            try:
                # Dùng engine='calamine' để đọc siêu nhanh nếu file quá nặng
                df = pd.read_excel(str(file_path), header=header_idx, usecols=[col_index], engine="calamine")
            except:
                # Dự phòng nếu máy chưa cài calamine
                df = pd.read_excel(str(file_path), header=header_idx, usecols=[col_index])
            
        real_col_name = df.columns[0]
        values = df[real_col_name].dropna().unique().tolist()
        return jsonify([str(v) for v in values])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/workflows/<workflow_id>/files", methods=["POST"])
def upload_workflow_file(workflow_id):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
        
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_dir = wf_dir / "input"
    input_dir.mkdir(parents=True, exist_ok=True)
    
    save_path = input_dir / file.filename
    file.save(str(save_path))
    
    return jsonify({"success": True, "name": file.filename})

@app.route("/api/workflows/<workflow_id>/files/<path:filename>", methods=["DELETE"])
def delete_workflow_file(workflow_id, filename):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_dir = wf_dir / "input"
    
    target_file = input_dir / filename
    if target_file.exists() and target_file.is_file() and target_file.name != "input.json":
        target_file.unlink()
        return jsonify({"success": True})
    return jsonify({"error": "File not found"}), 404

def open_file_in_os(filepath):
    try:
        if platform.system() == 'Windows':
            os.startfile(filepath)
        elif platform.system() == 'Darwin':
            subprocess.call(('open', filepath))
        else:
            subprocess.call(('xdg-open', filepath))
        return True
    except Exception as e:
        logger.error(f"Lỗi mở file: {e}")
        return False

@app.route("/api/workflows/<workflow_id>/files/<path:filename>/open", methods=["GET"])
def open_workflow_file(workflow_id, filename):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_dir = wf_dir / "input"
    
    target_file = input_dir / filename
    if target_file.exists() and target_file.is_file():
        open_file_in_os(str(target_file))
        return jsonify({"success": True})
    return jsonify({"error": "File not found"}), 404

@app.route("/api/workflows/<workflow_id>/files/<path:filename>/download", methods=["GET"])
def download_workflow_file(workflow_id, filename):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    input_dir = wf_dir / "input"
    
    target_file = input_dir / filename
    if target_file.exists() and target_file.is_file():
        as_attachment = request.args.get("download", "0") == "1"
        return send_file(target_file, as_attachment=as_attachment)
    return jsonify({"error": "File not found"}), 404

@app.route("/api/workflows/<workflow_id>/output-files", methods=["GET"])
def list_workflow_output_files(workflow_id):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    output_dir = wf_dir / "output"
    
    files = []
    if output_dir.exists():
        for f in output_dir.iterdir():
            if f.is_file():
                stat = f.stat()
                files.append({
                    "name": f.name,
                    "size": stat.st_size,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
    return jsonify(files)

@app.route("/api/workflows/<workflow_id>/output-files/<path:filename>/open", methods=["GET"])
def open_workflow_output_file(workflow_id, filename):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    output_dir = wf_dir / "output"
    
    target_file = output_dir / filename
    if target_file.exists() and target_file.is_file():
        open_file_in_os(str(target_file))
        return jsonify({"success": True})
    return jsonify({"error": "File not found"}), 404

@app.route("/api/workflows/<workflow_id>/output-files/<path:filename>/download", methods=["GET"])
def download_workflow_output_file(workflow_id, filename):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    output_dir = wf_dir / "output"
    
    target_file = output_dir / filename
    if target_file.exists() and target_file.is_file():
        as_attachment = request.args.get("download", "0") == "1"
        return send_file(target_file, as_attachment=as_attachment)
    return jsonify({"error": "File not found"}), 404

@app.route("/api/workflows/<workflow_id>/output-files/<path:filename>", methods=["DELETE"])
def delete_workflow_output_file(workflow_id, filename):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Not found"}), 404
        
    proj_dir = get_project_dir(wf["project_id"])
    wf_dir = proj_dir / f"wf_{slugify(wf['name'])}"
    output_dir = wf_dir / "output"
    
    target_file = output_dir / filename
    if target_file.exists() and target_file.is_file():
        target_file.unlink()
        return jsonify({"success": True})
    return jsonify({"error": "File not found"}), 404

@app.route("/api/workflows/<workflow_id>", methods=["DELETE"])
def delete_workflow(workflow_id):
    db = get_db()
    old_row = db.execute("SELECT name, project_id FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if old_row:
        import shutil
        proj_dir = get_project_dir(old_row["project_id"])
        wf_dir = proj_dir / f"wf_{slugify(old_row['name'])}"
        if wf_dir.exists():
            shutil.rmtree(wf_dir, ignore_errors=True)
            
    db.execute("DELETE FROM workflow WHERE id=?", (workflow_id,))
    db.commit()
    return "", 204


@app.route("/api/workflows/<workflow_id>/run", methods=["POST"])
def run_workflow(workflow_id):
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Workflow không tồn tại"}), 404
    if not wf["graph_json"]:
        return jsonify({"error": "Workflow chưa có graph. Thêm blocks trước."}), 400

    run_id = str(uuid.uuid4())
    now = now_iso()
    db.execute(
        "INSERT INTO workflow_run (id, workflow_id, project_id, status, started_at, triggered_by) VALUES (?,?,?,?,?,?)",
        (run_id, workflow_id, wf["project_id"], "running", now, "manual")
    )
    db.commit()

    stop_event = threading.Event()
    _active_runs[run_id] = stop_event
    # Đăng ký run_id vào workflow_id mapping
    _workflow_run_ids.setdefault(workflow_id, set()).add(run_id)
    log_fn = make_log_fn(run_id)

    start_time = datetime.now()
    t = threading.Thread(
        target=execute_workflow_thread,
        args=(run_id, wf["project_id"], workflow_id, wf["name"], wf["graph_json"], log_fn, stop_event),
        daemon=True
    )
    t.start()

    return jsonify({"run_id": run_id, "status": "started"})


@app.route("/api/workflows/<workflow_id>/stop", methods=["POST"])
def stop_workflow(workflow_id):
    stopped = False
    # Lấy tất cả run_id đang chạy của workflow này
    run_ids = list(_workflow_run_ids.get(workflow_id, set()))
    if not run_ids:
        # Fallback: dò DB để tìm run đang chạy
        try:
            with sqlite3.connect(str(DB_PATH)) as conn:
                rows = conn.execute(
                    "SELECT id FROM workflow_run WHERE workflow_id=? AND status='running'",
                    (workflow_id,)
                ).fetchall()
                run_ids = [r[0] for r in rows]
        except Exception:
            run_ids = list(_active_runs.keys())

    for run_id in run_ids:
        # Cập nhật DB ngay lập tức — không chờ thread xử lý
        try:
            with sqlite3.connect(str(DB_PATH)) as conn:
                conn.execute(
                    "UPDATE workflow_run SET status='stopped', finished_at=? WHERE id=? AND status='running'",
                    (datetime.now().isoformat(), run_id)
                )
        except Exception:
            pass

        stop_event = _active_runs.pop(run_id, None)
        if stop_event:
            stop_event.set()
            stopped = True
        # Force kill subprocess nếu có
        proc = _active_procs.pop(run_id, None)
        if proc:
            try:
                proc.kill()
            except Exception:
                pass

    _workflow_run_ids.pop(workflow_id, None)
    return jsonify({"stopped": stopped, "run_ids": run_ids})


@app.route("/api/workflows/<workflow_id>/runs", methods=["GET"])
def get_run_history(workflow_id):
    limit = request.args.get("limit", 20, type=int)
    db = get_db()
    rows = db.execute(
        "SELECT * FROM workflow_run WHERE workflow_id=? ORDER BY started_at DESC LIMIT ?",
        (workflow_id, limit)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/runs/<run_id>", methods=["GET"])
def get_run(run_id):
    db = get_db()
    row = db.execute("SELECT * FROM workflow_run WHERE id=?", (run_id,)).fetchone()
    if not row:
        return jsonify({"error": "Run không tồn tại"}), 404
    return jsonify(dict(row))


# ── SSE Log Streaming ─────────────────────────────────────

@app.route("/api/runs/<run_id>/logs/stream")
def stream_logs(run_id):
    """Server-Sent Events stream cho logs realtime"""
    import queue as Q
    q = Q.Queue(maxsize=1000)
    
    offset = request.args.get("offset", 0, type=int)
    
    # Nạp toàn bộ lịch sử log hiện có vào hàng đợi cho client mới
    cached_logs = _run_logs_cache.get(run_id, [])
    for msg in cached_logs[offset:]:
        try:
            q.put_nowait(msg)
        except Q.Full:
            pass

    if run_id not in _log_clients:
        _log_clients[run_id] = []
    _log_clients[run_id].append(q)

    def generate():
        try:
            yield f"data: {json.dumps({'type': 'connected', 'run_id': run_id})}\n\n"
            while True:
                try:
                    msg = q.get(timeout=30)
                    yield f"data: {msg}\n\n"
                except Q.Empty:
                    yield "data: {\"type\": \"ping\"}\n\n"
        except GeneratorExit:
            pass
        finally:
            if run_id in _log_clients:
                try:
                    _log_clients[run_id].remove(q)
                except ValueError:
                    pass

    from flask import Response
    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
    )


# ══════════════════════════════════════════════════════════════
#  API Routes — Schedules
# ══════════════════════════════════════════════════════════════

@app.route("/api/workflows/<workflow_id>/schedules", methods=["GET"])
def list_schedules(workflow_id):
    db = get_db()
    rows = db.execute("SELECT * FROM schedule WHERE workflow_id=?", (workflow_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/workflows/<workflow_id>/schedules", methods=["POST"])
def create_schedule(workflow_id):
    body = request.get_json(force=True)
    db = get_db()
    wf = db.execute("SELECT * FROM workflow WHERE id=?", (workflow_id,)).fetchone()
    if not wf:
        return jsonify({"error": "Workflow không tồn tại"}), 404

    cron_expr = body.get("cron_expr", "0 8 * * *")
    sid = str(uuid.uuid4())
    next_run = get_next_run(cron_expr)
    enabled = body.get("enabled", True)

    db.execute(
        "INSERT INTO schedule (id, workflow_id, cron_expr, label, enabled, created_at, next_run_at) VALUES (?,?,?,?,?,?,?)",
        (sid, workflow_id, cron_expr, body.get("label") or cron_expr, 1 if enabled else 0, now_iso(), next_run)
    )
    db.commit()

    if enabled:
        try:
            scheduler.add_job(
                _scheduler_trigger, "cron", id=sid,
                kwargs={"workflow_id": workflow_id, "project_id": wf["project_id"], "schedule_id": sid},
                replace_existing=True, **_cron_kwargs(cron_expr),
            )
        except Exception as e:
            logger.warning(f"Lỗi add job scheduler: {e}")

    row = db.execute("SELECT * FROM schedule WHERE id=?", (sid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/schedules/<schedule_id>", methods=["PUT"])
def update_schedule(schedule_id):
    body = request.get_json(force=True)
    db = get_db()
    sched = db.execute("SELECT * FROM schedule WHERE id=?", (schedule_id,)).fetchone()
    if not sched:
        return jsonify({"error": "Schedule không tồn tại"}), 404

    fields = {k: body[k] for k in ["cron_expr", "label", "enabled"] if k in body}
    if "cron_expr" in fields:
        fields["next_run_at"] = get_next_run(fields["cron_expr"])
    if fields:
        sets = ", ".join(f"{k}=?" for k in fields)
        db.execute(f"UPDATE schedule SET {sets} WHERE id=?", (*fields.values(), schedule_id))
        db.commit()

    sched = db.execute("SELECT s.*, w.project_id FROM schedule s JOIN workflow w ON s.workflow_id=w.id WHERE s.id=?", (schedule_id,)).fetchone()
    if sched["enabled"]:
        try:
            scheduler.add_job(
                _scheduler_trigger, "cron", id=schedule_id,
                kwargs={"workflow_id": sched["workflow_id"], "project_id": sched["project_id"], "schedule_id": schedule_id},
                replace_existing=True, **_cron_kwargs(sched["cron_expr"]),
            )
        except Exception as e:
            logger.warning(f"Lỗi update scheduler job: {e}")
    else:
        try:
            scheduler.remove_job(schedule_id)
        except Exception:
            pass

    return jsonify(dict(db.execute("SELECT * FROM schedule WHERE id=?", (schedule_id,)).fetchone()))


@app.route("/api/schedules/<schedule_id>/toggle", methods=["PATCH"])
def toggle_schedule(schedule_id):
    db = get_db()
    sched = db.execute(
        "SELECT s.*, w.project_id FROM schedule s JOIN workflow w ON s.workflow_id=w.id WHERE s.id=?",
        (schedule_id,)
    ).fetchone()
    if not sched:
        return jsonify({"error": "Schedule không tồn tại"}), 404

    new_enabled = not bool(sched["enabled"])
    next_run = get_next_run(sched["cron_expr"]) if new_enabled else None
    db.execute(
        "UPDATE schedule SET enabled=?, next_run_at=? WHERE id=?",
        (1 if new_enabled else 0, next_run, schedule_id)
    )
    db.commit()

    if new_enabled:
        try:
            scheduler.add_job(
                _scheduler_trigger, "cron", id=schedule_id,
                kwargs={"workflow_id": sched["workflow_id"], "project_id": sched["project_id"], "schedule_id": schedule_id},
                replace_existing=True, **_cron_kwargs(sched["cron_expr"]),
            )
        except Exception as e:
            logger.warning(f"Lỗi resume scheduler: {e}")
    else:
        try:
            scheduler.remove_job(schedule_id)
        except Exception:
            pass

    return jsonify({"id": schedule_id, "enabled": new_enabled, "next_run_at": next_run})


@app.route("/api/schedules/<schedule_id>", methods=["DELETE"])
def delete_schedule(schedule_id):
    db = get_db()
    try:
        scheduler.remove_job(schedule_id)
    except Exception:
        pass
    db.execute("DELETE FROM schedule WHERE id=?", (schedule_id,))
    db.commit()
    return "", 204


# ══════════════════════════════════════════════════════════════
#  API Routes — Users
# ══════════════════════════════════════════════════════════════

@app.route("/api/users", methods=["GET"])
def list_users():
    db = get_db()
    rows = db.execute("SELECT * FROM user ORDER BY created_at ASC").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/users", methods=["POST"])
def create_user():
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Thiếu tên người dùng"}), 400
    uid = str(uuid.uuid4())
    now = now_iso()
    db = get_db()
    try:
        db.execute("INSERT INTO user (id, name, created_at) VALUES (?,?,?)", (uid, name, now))
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": f"Tên '{name}' đã tồn tại"}), 409
    row = db.execute("SELECT * FROM user WHERE id=?", (uid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/users/<user_id>", methods=["DELETE"])
def delete_user(user_id):
    db = get_db()
    # Cascade-delete projects (và qua FK: workflows, schedules, runs)
    project_ids = [r[0] for r in db.execute(
        "SELECT id FROM project WHERE user_id=?", (user_id,)
    ).fetchall()]
    for pid in project_ids:
        delete_venv_sync(pid)
    db.execute("DELETE FROM project WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM user WHERE id=?", (user_id,))
    db.commit()
    return "", 204


@app.route("/api/users/<user_id>/activate", methods=["POST"])
def activate_user(user_id):
    """Chuyển sang user mới: set is_active, dừng all jobs cũ, load lại schedules của user này"""
    db = get_db()
    user = db.execute("SELECT * FROM user WHERE id=?", (user_id,)).fetchone()
    if not user:
        return jsonify({"error": "Người dùng không tồn tại"}), 404

    # Cập nhật is_active: reset tất cả, set user này = 1
    db.execute("UPDATE user SET is_active=0")
    db.execute("UPDATE user SET is_active=1 WHERE id=?", (user_id,))
    db.commit()

    # Xóa tất cả jobs hiện tại khỏi scheduler
    for job in scheduler.get_jobs():
        try:
            job.remove()
        except Exception:
            pass

    # Load lại chỉ schedules của user này
    rows = db.execute("""
        SELECT s.*, w.project_id FROM schedule s
        JOIN workflow w ON s.workflow_id = w.id
        JOIN project p ON w.project_id = p.id
        WHERE s.enabled = 1 AND p.user_id = ?
    """, (user_id,)).fetchall()

    loaded = 0
    for row in rows:
        try:
            scheduler.add_job(
                _scheduler_trigger,
                "cron",
                id=row["id"],
                kwargs={"workflow_id": row["workflow_id"], "project_id": row["project_id"], "schedule_id": row["id"]},
                **_cron_kwargs(row["cron_expr"]),
                replace_existing=True,
            )
            loaded += 1
        except Exception as e:
            logger.warning(f"  ⚠ Lỗi load schedule {row['id']}: {e}")

    logger.info(f"👤 Activated user '{user['name']}' — loaded {loaded} schedules")
    return jsonify({"status": "ok", "user": dict(user), "schedules_loaded": loaded})


@app.route("/api/users/<user_id>/stats", methods=["GET"])
def get_user_stats(user_id):
    """Trả về thống kê của user: số projects, số workflows"""
    db = get_db()
    project_count = db.execute(
        "SELECT COUNT(*) as c FROM project WHERE user_id=?", (user_id,)
    ).fetchone()["c"]
    workflow_count = db.execute(
        "SELECT COUNT(*) as c FROM workflow WHERE project_id IN (SELECT id FROM project WHERE user_id=?)",
        (user_id,)
    ).fetchone()["c"]
    schedule_count = db.execute(
        """SELECT COUNT(*) as c FROM schedule s
           JOIN workflow w ON s.workflow_id=w.id
           JOIN project p ON w.project_id=p.id
           WHERE p.user_id=?""",
        (user_id,)
    ).fetchone()["c"]
    return jsonify({
        "user_id": user_id,
        "project_count": project_count,
        "workflow_count": workflow_count,
        "schedule_count": schedule_count,
    })


# ══════════════════════════════════════════════════════════════
#  Health & Misc
# ══════════════════════════════════════════════════════════════


@app.route("/health")
def health():
    return jsonify({"status": "ok", "service": "PyFlow Studio Backend", "version": "1.0.0"})


@app.route("/api/scheduler/jobs")
def get_scheduler_jobs():
    jobs = []
    for job in scheduler.get_jobs():
        jobs.append({
            "id": job.id,
            "next_run_time": job.next_run_time.isoformat() if job.next_run_time else None,
        })
    return jsonify({"jobs": jobs})


# ══════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 45)
    print("  PyFlow Studio -- Backend v1.0")
    print("  http://localhost:8000")
    print("=" * 45)

    # Init DB
    init_db()

    # Cleanup stale runs từ phiên trước
    cleanup_stale_runs()

    # Start Scheduler
    scheduler.start()
    logger.info("🟢 APScheduler đã khởi động")

    # Load schedules từ DB
    load_schedules_from_db()

    # Start Flask
    app.run(host="0.0.0.0", port=8000, debug=False, threaded=True, use_reloader=False)
