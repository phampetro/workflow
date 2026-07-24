import os
import sys
import json
import uuid
import sqlite3
import logging
import urllib.request
import urllib.error
import urllib.parse
import smtplib
from email.message import EmailMessage
import asyncio
import subprocess
import threading
import unicodedata
import re
from datetime import datetime
from pathlib import Path
import platform
import zipfile
import shutil
import io
import pandas as pd

from services import venv_manager

DATA_DIR = venv_manager.DATA_DIR
WORKFLOW_DB = DATA_DIR / "pyflow.db"

_active_runs = {}
_active_procs = {}
_workflow_run_ids = {}

_active_listeners = {}
_active_browser_profiles = {}

def kill_run(run_id):
    """Force kill tất cả processes liên quan đến một run"""
    # Kill subproc nếu đang chạy
    if run_id in _active_procs:
        proc = _active_procs.pop(run_id, None)
        if proc:
            try:
                proc.kill()
            except Exception:
                pass
    # Set cờ dừng để vòng lặp workflow không chạy tiếp block kế
    ev = _active_runs.pop(run_id, None)
    if ev is not None:
        try:
            ev.set()
        except Exception:
            pass


def stop_workflow_by_id(run_id):
    """Stop a specific workflow run by killing its processes"""
    kill_run(run_id)


def stop_all_runs_for_workflow(workflow_id):
    """Stop all active runs for a workflow"""
    # Đọc set qua .get(...) rồi wrap list() ngay - trước đây if-in-then-index có 2 bước
    # không atomic: _finish_run ở thread khác có thể pop key giữa 2 bước, gây KeyError
    # crash endpoint /stop.
    for run_id in list(_workflow_run_ids.get(workflow_id, set())):
        kill_run(run_id)

def now_iso() -> str:
    return datetime.now().astimezone().isoformat()

def slugify(text: str) -> str:
    if not text:
        return "untitled"
    import unicodedata
    import re
    text = unicodedata.normalize('NFKD', str(text)).encode('ascii', 'ignore').decode('utf-8')
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    text = re.sub(r'[-\s_]+', '_', text)
    return text

def get_project_dir(project_id: str) -> Path:
    db_path = venv_manager.DATA_DIR / "pyflow.db"
    with sqlite3.connect(str(db_path), timeout=5) as conn:
        row = conn.execute("SELECT name FROM project WHERE id=?", (project_id,)).fetchone()
        name = row[0] if row else "unknown"
    return DATA_DIR / f"pj_{slugify(name)}"

def get_saved_db_connection(connection_id: str) -> dict | None:
    if not connection_id:
        return None
    db_path = venv_manager.DATA_DIR / "pyflow.db"
    with sqlite3.connect(str(db_path), timeout=5) as conn:
        row = conn.execute(
            "SELECT db_type, host, port, username, password, dbname FROM db_connection WHERE id=?",
            (connection_id,)
        ).fetchone()
    if not row:
        return None
    return {
        "db_type": row[0], "host": row[1], "port": row[2],
        "username": row[3], "password": row[4], "dbname": row[5],
    }

def build_conn_str(db_config: dict) -> str:
    db_type = db_config.get("db_type") or "sqlserver"
    host = db_config.get("host") or ""
    port = db_config.get("port") or ""
    dbname = db_config.get("dbname") or ""
    user = db_config.get("username") or ""
    pwd = db_config.get("password") or ""

    user_enc = urllib.parse.quote_plus(str(user)) if user else ""
    pwd_enc = urllib.parse.quote_plus(str(pwd)) if pwd else ""
    port_str = (":" + str(port)) if port else ""

    if db_type == "postgresql":
        return "postgresql://" + user_enc + ":" + pwd_enc + "@" + host + port_str + "/" + dbname
    elif db_type == "mysql":
        return "mysql+pymysql://" + user_enc + ":" + pwd_enc + "@" + host + port_str + "/" + dbname
    elif db_type == "sqlite":
        return "sqlite:///" + dbname
    else:
        return "mssql+pyodbc://" + user_enc + ":" + pwd_enc + "@" + host + port_str + "/" + dbname + "?driver=ODBC+Driver+17+for+SQL+Server&TrustServerCertificate=yes"

def get_venv_dir(project_id: str) -> Path:
    return venv_manager.get_venv_path(project_id)

def get_python_path(project_id: str) -> str:
    return venv_manager.get_python_path(project_id)

def get_pip_path(project_id: str) -> str:
    return venv_manager.get_pip_path(project_id)

def venv_exists(project_id: str) -> bool:
    return venv_manager.venv_exists(project_id)

def _stop_telegram_listener_sync(workflow_id: str, log_fn=None):
    """Dừng Telegram Listener từ thread thực thi workflow (đồng bộ, không có event loop riêng)."""
    try:
        from services.telegram_listener import _stop_events, _active_listeners
        if workflow_id in _stop_events:
            _stop_events[workflow_id].set()
        task = _active_listeners.get(workflow_id)
        if task:
            try:
                # Lấy đúng loop mà task listener đang chạy trên đó (thread riêng),
                # KHÔNG dùng asyncio.get_running_loop() vì hàm này được gọi từ một
                # thread đồng bộ không có loop nào đang chạy cả.
                task.get_loop().call_soon_threadsafe(task.cancel)
            except RuntimeError:
                # Loop của task đã đóng - stop_event.set() ở trên vẫn đủ để vòng
                # lặp long-polling tự thoát trong lần lặp kế tiếp (nếu còn sống).
                pass
    except Exception as e:
        if log_fn:
            log_fn("system", "warning", f"⚠ Không tắt được listener: {e}")


def _set_workflow_listener_flag(workflow_id: str, on: bool):
    """Cập nhật cột workflow.listener_on trong DB (sync, chạy trong thread executor)."""
    try:
        with sqlite3.connect(str(WORKFLOW_DB), timeout=5) as conn:
            conn.execute(
                "UPDATE workflow SET listener_on=? WHERE id=?",
                (1 if on else 0, workflow_id)
            )
    except Exception:
        pass


def create_venv_sync(project_id: str) -> dict:
    proj_dir = get_project_dir(project_id)
    proj_dir.mkdir(parents=True, exist_ok=True)
    venv_path = get_venv_dir(project_id)
    venv_path.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [sys.executable, "-m", "venv", str(venv_path)],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0:
        raise RuntimeError(f"Không tạo được venv: {result.stderr}")
    return {"path": str(venv_path), "python": get_python_path(project_id)}


def install_pkg_sync(project_id: str, package: str) -> dict:
    if not venv_exists(project_id):
        create_venv_sync(project_id)
    pip = get_pip_path(project_id)
    result = subprocess.run([pip, "install", package, "--quiet"], capture_output=True, text=True, timeout=300)
    if result.returncode != 0:
        raise RuntimeError(f"Cài thất bại: {result.stderr[:500]}")
    return {"package": package, "status": "installed"}


def uninstall_pkg_sync(project_id: str, package: str) -> dict:
    pip = get_pip_path(project_id)
    subprocess.run([pip, "uninstall", package, "-y"], capture_output=True, text=True, timeout=60)
    return {"package": package, "status": "uninstalled"}

def ensure_packages(project_id: str, packages: list, log_fn=None, bid=None, label="", stop_event=None):
    if not packages: return
    if not venv_exists(project_id):
        create_venv_sync(project_id)
    
    pip = get_pip_path(project_id)
    result = subprocess.run([pip, "freeze"], capture_output=True, text=True, timeout=30)
    installed = result.stdout.lower()
    
    missing = []
    for pkg in packages:
        pkg_name = pkg.lower().split('==')[0].split('>')[0].split('<')[0]
        if f"{pkg_name}==" not in installed and f"{pkg_name} @" not in installed:
            missing.append(pkg)
            
    if missing:
        if log_fn:
            log_fn(bid, "info", f"📦 [{label}] Đang tải & cài đặt: {', '.join(missing)}...")
        
        import time
        res = subprocess.Popen(
            [pip, "install"] + missing,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            bufsize=1,
            text=True,
            encoding="utf-8",
            errors="replace"
        )
        
        # Stream output với kiểm tra stop và timeout
        last_check = time.time()
        check_interval = 1.0  # Check mỗi giây
        while True:
            # Check stop event
            if stop_event and stop_event.is_set():
                res.kill()
                if log_fn:
                    log_fn(bid, "warning", f"⏹ Cài đặt bị dừng bởi người dùng")
                raise RuntimeError("Installation stopped by user")
            
            # Non-blocking read with timeout
            line = res.stdout.readline()
            if line:
                line = line.strip()
                if line and log_fn:
                    log_fn(bid, "info", f"      [pip] {line}")
                last_check = time.time()
            else:
                # No more output, check if process ended
                if res.poll() is not None:
                    break
                # Small sleep to avoid busy loop
                time.sleep(0.1)
                
                # Check timeout (5 minutes max for installation)
                if time.time() - last_check > 300:
                    res.kill()
                    if log_fn:
                        log_fn(bid, "error", f"❌ Cài đặt timeout sau 5 phút")
                    raise RuntimeError(f"Cài đặt {missing} timeout")
        
        if res.returncode != 0:
            if log_fn:
                log_fn(bid, "error", f"❌ Cài đặt thất bại mã {res.returncode}")
            raise RuntimeError(f"Không thể cài đặt {missing}")

def list_pkgs_sync(project_id: str) -> list:
    if not venv_exists(project_id):
        return []
    pip = get_pip_path(project_id)
    result = subprocess.run([pip, "list", "--format=json"], capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        return []
    try:
        return [{"name": p["name"], "version": p["version"]} for p in json.loads(result.stdout)]
    except Exception:
        return []


def delete_venv_sync(project_id: str):
    import shutil
    venv_path = get_venv_dir(project_id)
    if venv_path.exists():
        shutil.rmtree(venv_path, ignore_errors=True)


# ══════════════════════════════════════════════════════════════
#  Execution Engine
# ══════════════════════════════════════════════════════════════

RUNNER_TEMPLATE = '''
import sys, json, traceback, os

workflow_id = {workflow_id!r}
block_id = {block_id!r}
OUTPUT_DIR = {output_dir!r}
INPUT_DIR = {input_dir!r}

input_data = None
try:
    _raw = sys.stdin.read().strip()
    if _raw:
        input_data = json.loads(_raw)
except Exception:
    pass

output_data = None

try:
{user_code}
except Exception as _e:
    print("[ERROR] " + traceback.format_exc(), file=sys.stderr)
    sys.exit(1)

try:
    _out = json.dumps(output_data, ensure_ascii=False, default=str)
except Exception:
    _out = json.dumps(str(output_data))
print("__OUTPUT__:" + _out)
'''


def indent_code(code: str, spaces: int = 4) -> str:
    return "\n".join(" " * spaces + line for line in code.splitlines())


def topological_sort(nodes: list, edges: list) -> list:
    adj = {n["id"]: [] for n in nodes}
    in_deg = {n["id"]: 0 for n in nodes}
    for edge in edges:
        src, tgt = edge.get("source"), edge.get("target")
        if src in adj and tgt in adj:
            adj[src].append(tgt)
            in_deg[tgt] += 1
    queue = [nid for nid, d in in_deg.items() if d == 0]
    ordered = []
    while queue:
        nid = queue.pop(0)
        ordered.append(nid)
        for nxt in adj.get(nid, []):
            in_deg[nxt] -= 1
            if in_deg[nxt] == 0:
                queue.append(nxt)
    node_map = {n["id"]: n for n in nodes}
    return [node_map[nid] for nid in ordered if nid in node_map]


def get_workflow_dir(project_id: str, workflow_id: str) -> Path:
    with sqlite3.connect(str(WORKFLOW_DB), timeout=5) as conn:
        row = conn.execute("SELECT name FROM workflow WHERE id=?", (workflow_id,)).fetchone()
        name = row[0] if row else "unknown"
    return get_project_dir(project_id) / f"wf_{slugify(name)}"

def run_python_block_sync(project_id, block_id, workflow_id, code, input_data, timeout=60, label=None, log_fn=None, input_dir=None, stop_event=None):
    """Chạy 1 block Python synchronously, có thể bị ngắt bởi stop_event"""
    if not venv_exists(project_id):
        create_venv_sync(project_id)

    python_exe = get_python_path(project_id)
    wf_dir = get_workflow_dir(project_id, workflow_id)
    wf_dir.mkdir(parents=True, exist_ok=True)
    block_dir = wf_dir / slugify(label or block_id)
    block_dir.mkdir(parents=True, exist_ok=True)
    output_dir = wf_dir / "output"
    output_dir.mkdir(exist_ok=True)

    block_path = block_dir / "main.py"
    wrapped = RUNNER_TEMPLATE.format(
        workflow_id=workflow_id,
        block_id=block_id,
        output_dir=str(output_dir).replace('\\', '/'),
        input_dir=str(input_dir).replace('\\', '/') if input_dir else '',
        user_code=indent_code(code),
    )
    block_path.write_text(wrapped, encoding="utf-8")

    input_json = json.dumps(input_data, ensure_ascii=False, default=str)

    if log_fn:
        log_fn(block_id, "info", f"▶  Chạy block [{label or block_id}]")

    start = datetime.now()
    run_id_for_proc = None
    # Tìm run_id tương ứng dựa vào stop_event
    for rid, ev in list(_active_runs.items()):
        if ev is stop_event:
            run_id_for_proc = rid
            break

    try:
        proc = subprocess.Popen(
            [python_exe, "-u", "-X", "utf8", str(block_path)],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,  # Line buffered
            env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )

        # Đăng ký proc để có thể force-kill
        if run_id_for_proc:
            _active_procs[run_id_for_proc] = proc

        def write_stdin():
            try:
                proc.stdin.write(input_json.encode("utf-8"))
                proc.stdin.close()
            except Exception:
                pass

        t_in = threading.Thread(target=write_stdin, daemon=True)
        t_in.start()

        output_lines = []
        stderr_lines = []

        def read_stdout():
            try:
                for line in proc.stdout:
                    line_str = line.decode("utf-8", errors="replace").rstrip("\r\n")
                    if line_str.startswith("__OUTPUT__:"):
                        output_lines.append(line_str)
                    elif line_str.strip() and log_fn:
                        log_fn(block_id, "info", f"   {line_str}")
            except Exception:
                pass

        def read_stderr():
            try:
                for line in proc.stderr:
                    line_str = line.decode("utf-8", errors="replace").rstrip("\r\n")
                    if line_str.strip():
                        stderr_lines.append(line_str)
                        if log_fn:
                            log_fn(block_id, "error", f"   ⚠ {line_str}")
            except Exception:
                pass

        t_out = threading.Thread(target=read_stdout, daemon=True)
        t_err = threading.Thread(target=read_stderr, daemon=True)
        t_out.start()
        t_err.start()

        # Chờ proc và check stop_event định kỳ
        deadline = datetime.now().timestamp() + timeout
        while True:
            # Chờ tối đa 0.5s mỗi lần
            try:
                proc.wait(timeout=0.5)
                break  # proc kết thúc
            except subprocess.TimeoutExpired:
                pass

            # Kiểm tra stop_event
            if stop_event and stop_event.is_set():
                import signal
                try:
                    proc.kill()
                except Exception:
                    pass
                duration = int((datetime.now() - start).total_seconds() * 1000)
                if log_fn:
                    log_fn(block_id, "warning", f"⏹ Block đã bị dừng ({duration}ms)")
                if run_id_for_proc:
                    _active_procs.pop(run_id_for_proc, None)
                return False, None, "stopped", duration

            # Kiểm tra timeout
            if datetime.now().timestamp() > deadline:
                try:
                    proc.kill()
                except Exception:
                    pass
                duration = int((datetime.now() - start).total_seconds() * 1000)
                msg = f"⏰ Timeout sau {timeout}s"
                if log_fn:
                    log_fn(block_id, "error", msg)
                if run_id_for_proc:
                    _active_procs.pop(run_id_for_proc, None)
                return False, None, msg, duration

        if run_id_for_proc:
            _active_procs.pop(run_id_for_proc, None)

        duration = int((datetime.now() - start).total_seconds() * 1000)
        output_data = None
        
        # Ensure threads have finished reading
        t_out.join(timeout=1)
        t_err.join(timeout=1)

        for line in output_lines:
            try:
                output_data = json.loads(line[len("__OUTPUT__:"):])
            except Exception:
                output_data = line[len("__OUTPUT__:"):]

        if proc.returncode != 0:
            err = "\n".join(stderr_lines).strip() or "Unknown error"
            if log_fn:
                log_fn(block_id, "error", f"✗ Block thất bại ({duration}ms)")
            return False, None, err, duration

        if log_fn:
            log_fn(block_id, "success", f"✓ Block hoàn thành ({duration}ms)")
        return True, output_data, None, duration

    except Exception as e:
        if run_id_for_proc:
            _active_procs.pop(run_id_for_proc, None)
        duration = int((datetime.now() - start).total_seconds() * 1000)
        if log_fn:
            log_fn(block_id, "error", f"✗ Lỗi: {e}")
        return False, None, str(e), duration


def execute_workflow_thread(run_id, project_id, workflow_id, workflow_name, graph_json, log_fn, stop_event):
    """Chạy toàn bộ workflow trong thread riêng"""
    import time
    time.sleep(0.5) # Đợi client SSE kết nối trước khi chạy nhanh
    start = datetime.now()

    # Đăng ký run để cơ chế force-kill (kill_run / stop_all_runs_for_workflow)
    # tìm được stop_event và subprocess tương ứng. _finish_run sẽ dọn dẹp.
    _active_runs[run_id] = stop_event
    _workflow_run_ids.setdefault(workflow_id, set()).add(run_id)

    try:
        graph = json.loads(graph_json)
    except Exception as e:
        _finish_run(run_id, "error", start, error=str(e))
        return

    try:
        proj_dir = get_project_dir(project_id)
        wf_dir = proj_dir / f"wf_{slugify(workflow_name)}"
        input_dir = wf_dir / "input"
        
        # Thư mục chứa cấu hình và phiên đăng nhập (Profile) cho riêng lượt chạy này
        browser_profile_dir = wf_dir / "runs" / run_id / "browser_profile"
        _active_browser_profiles[run_id] = browser_profile_dir
        input_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        # Lỗi trước vòng try chính (DB lock, PermissionError...) không được
        # để run kẹt ở trạng thái RUNNING vĩnh viễn
        _finish_run(run_id, "error", start, error=str(e))
        if log_fn:
            log_fn("system", "error", f"❌ Lỗi chuẩn bị thư mục workflow: {e}")
        return

    try:
        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])
        
        nodes_dict = {n["id"]: n for n in nodes}
        edges_from = {}
        for e in edges:
            edges_from.setdefault(e["source"], []).append(e)

        listener_nodes = [n for n in nodes if n.get("data", {}).get("type") == "telegram_listener" and "_initial_input" in n.get("data", {})]
        
        if listener_nodes:
            start_nodes = listener_nodes
            initial_input = listener_nodes[0]["data"]["_initial_input"]
        else:
            start_nodes = [n for n in nodes if n.get("data", {}).get("type") == "start"]
            initial_input = None

        if not start_nodes:
            if log_fn:
                log_fn("system", "error", "❌ Không tìm thấy khối Bắt đầu!")
            _finish_run(run_id, "error", start, error="Missing Start block")
            return

        from collections import deque
        queue = deque()
        # Enqueue: (node_id, input_data)
        
        error_trigger_nodes = [n for n in nodes if n.get("data", {}).get("type") == "error_trigger"]
        in_error_mode = False
        
        def handle_workflow_error(error_msg, failed_bid=None, failed_label=None):
            nonlocal in_error_mode, queue
            if in_error_mode:
                _finish_run(run_id, "error", start, error=error_msg)
                return True
            
            if error_trigger_nodes:
                if log_fn:
                    log_fn("system", "warning", f"⚠️ Phát hiện lỗi{' tại ['+failed_label+']' if failed_label else ''}. Đang chuyển hướng sang khối Bắt Lỗi toàn cục...")
                queue.clear()
                in_error_mode = True
                
                error_payload = {
                    "status": "error",
                    "error_detail": error_msg,
                    "failed_block": failed_label,
                    "failed_block_id": failed_bid
                }
                for et_node in error_trigger_nodes:
                    queue.append((et_node["id"], error_payload))
                return False
            else:
                _finish_run(run_id, "error", start, error=error_msg)
                return True
        queue.append((start_nodes[0]["id"], initial_input))

        if log_fn:
            log_fn("system", "info", f"🚀 Bắt đầu workflow (Dynamic Routing)")

        import collections
        final_status = "success"
        run_counts = collections.Counter()
        loop_states = {}

        while queue:
            if stop_event and stop_event.is_set():
                log_fn("system", "warning", "⏹ Đã dừng bởi người dùng")
                final_status = "stopped"
                break

            node_id, current_input = queue.popleft()
            if run_counts[node_id] > 2000:
                if log_fn:
                    log_fn("system", "error", f"❌ Phát hiện lặp vô hạn ở Node {node_id} (>2000 lần). Dừng luồng.")
                final_status = "error"
                break
            run_counts[node_id] += 1

            node = nodes_dict.get(node_id)
            if not node:
                continue

            bdata = node.get("data", {})
            
            if "workflow_env" not in locals():
                workflow_env = {}
                try:
                    input_file = input_dir / "input.json"
                    if input_file.exists():
                        with open(input_file, "r", encoding="utf-8") as f:
                            content_str = f.read()
                            pattern = r'("(?:\\.|[^"\\])*")|(//.*)|(/\*[\s\S]*?\*/)'
                            cleaned = re.sub(pattern, lambda m: m.group(1) if m.group(1) else '', content_str)
                            cleaned = re.sub(r',(\s*[}\]])', r'\1', cleaned)
                            workflow_env = json.loads(cleaned)
                except Exception:
                    pass
                    
            def interpolate(val):
                if not isinstance(val, str):
                    return val

                ctx = dict(workflow_env)
                if isinstance(current_input, dict):
                    ctx.update(current_input)
                ctx["input_data"] = current_input

                if val in ctx:
                    return str(ctx[val])

                # Dùng re.sub 1 lượt để giá trị được thay KHÔNG bị scan tiếp - nếu không,
                # value của biến k1 chứa chuỗi "{{k2}}" sẽ bị thay tiếp ở vòng sau,
                # gây rò rỉ dữ liệu (VD sender_name Telegram do người ngoài đặt).
                def _sub(m):
                    key = m.group(1)
                    return str(ctx[key]) if key in ctx else m.group(0)
                return re.sub(r"\{\{(\w+)\}\}", _sub, val)

            def interpolate_deep(val, current_key=None):
                # Nội suy đệ quy vào dict/list lồng nhau (conditions, attachments,
                # danh sách lệnh telegram...) — không chỉ field string cấp cao nhất
                if isinstance(val, str):
                    # Bỏ qua nội suy đối với các trường dùng để điền "tên biến" và "code" python
                    if current_key in ("condVariable", "key_name", "code"):
                        return val
                    return interpolate(val)
                if isinstance(val, dict):
                    return {k: interpolate_deep(v, k) for k, v in val.items()}
                if isinstance(val, list):
                    return [interpolate_deep(item, current_key) for item in val]
                return val

            btype = bdata.get("type", "python")
            bid = node["id"]
            label = bdata.get("label", bid)

            # Nội suy các biến môi trường (đệ quy vào cả cấu trúc lồng nhau)
            # Ngoại trừ "steps" của khối Browser để giữ nguyên {{biến}} cho browser_executor tự nội suy động.
            bdata_interpolated = {}
            for k, v in bdata.items():
                if btype == "browser" and k == "steps":
                    bdata_interpolated[k] = v
                else:
                    bdata_interpolated[k] = interpolate_deep(v, k)
            bdata = bdata_interpolated
            label = bdata.get("label", bid)

            # Execution flags for branching
            continue_branch = True
            cond_branch_taken = None # 'true' or 'false'

            if btype == "start":
                if log_fn:
                    log_fn(bid, "info", f"▶  [Start] {label}")
            elif btype == "end":
                if log_fn:
                    log_fn(bid, "info", f"🏁 [End] {label}")
                break
            elif btype == "delay":
                delay_sec = float(bdata.get("delaySeconds", 3))
                if log_fn:
                    log_fn(bid, "info", f"⏳ [Delay] {label} - Đang chờ {delay_sec} giây...")
                
                # Chờ có kiểm tra stop_event định kỳ (mỗi 0.5s)
                import time
                waited = 0.0
                check_interval = 0.5
                while waited < delay_sec:
                    if stop_event and stop_event.is_set():
                        if log_fn:
                            log_fn(bid, "warning", f"⏹ Delay bị dừng sau {int(waited)}s")
                        final_status = "stopped"
                        break
                    sleep_time = min(check_interval, delay_sec - waited)
                    time.sleep(sleep_time)
                    waited += sleep_time
                else:
                    if log_fn:
                        log_fn(bid, "success", f"✅ Đã chờ xong {delay_sec} giây.")
            elif btype == "queue":
                # Khối "Xếp hàng": 1 vào - 1 ra, không xử lý gì, không đổi biến.
                # Cơ chế "lấy số chờ tới lượt": nếu trong hàng đợi còn khối THƯỜNG
                # (không phải queue) đang chờ chạy → tự lùi xuống cuối hàng, nhường
                # cho các nhánh song song chạy hết. Chỉ khi không còn khối thường nào
                # chờ mới đi tiếp → khối phía sau luôn chạy SAU CÙNG và đúng 1 lần.
                # (Điều kiện loại trừ queue giúp nhiều khối queue không "nhường" nhau vô tận.)
                non_queue_pending = any(
                    ((nodes_dict.get(nid) or {}).get("data") or {}).get("type") != "queue"
                    for nid, _ in queue
                )
                if non_queue_pending:
                    run_counts[node_id] -= 1  # lần lùi hàng không tính là 1 lần chạy thật
                    queue.append((node_id, current_input))
                    continue
                if log_fn:
                    log_fn(bid, "info", f"⏳ [Xếp hàng] {label} - Đã tới lượt, chạy tiếp")
            elif btype == "telegram_listener":
                if "_initial_input" in bdata:
                    if log_fn:
                        log_fn(bid, "info", f"🎧 [Telegram Listener] {label} - Đã nhận tin nhắn và chạy workflow")
                    current_input = bdata["_initial_input"]
                    if isinstance(current_input, dict):
                        for field_key, bdata_key in (
                            ("chat_id", "telegramListenerChatIdVarName"),
                            ("message_id", "telegramListenerMessageIdVarName"),
                            ("text", "telegramListenerTextVarName"),
                            ("sender_name", "telegramListenerSenderNameVarName"),
                        ):
                            var_name = bdata.get(bdata_key, "").strip()
                            if var_name:
                                workflow_env[var_name] = current_input.get(field_key)
                else:
                    # Manual run (bấm nút Chạy/Start) → bật Listener, workflow giữ RUNNING chờ tin nhắn
                    if log_fn:
                        log_fn(bid, "info", f"🎧 [Telegram Listener] {label} - Đang bật Listener để chờ tin nhắn...")

                    # Tự động bật Listener (nếu chưa bật) - chạy trong thread riêng có event loop
                    try:
                        import threading as _threading_listener
                        from services.telegram_listener import (
                            start_telegram_listener,
                            is_listener_running,
                            get_listener_config,
                        )

                        # Tính token + commands từ bdata (đã nội suy biến {{key}}/input.json)
                        tg_token = bdata.get("telegramListenerToken") or bdata.get("telegramBotToken", "")

                        raw_commands = bdata.get("telegramListenerCommands") or bdata.get("telegramCommands", "")
                        tg_commands = []
                        if isinstance(raw_commands, list):
                            for c in raw_commands:
                                if isinstance(c, dict):
                                    tg_commands.append({
                                        "command": (c.get("command") or "").strip(),
                                        "description": c.get("description", ""),
                                        "reply": c.get("reply", ""),
                                        "runWorkflow": bool(c.get("runWorkflow", False)),
                                    })
                                else:
                                    tg_commands.append({"command": str(c).strip(), "reply": "", "runWorkflow": True})
                        else:
                            tg_commands = [
                                {"command": c.strip(), "reply": "", "runWorkflow": True}
                                for c in str(raw_commands).split(",") if c.strip()
                            ]

                        # Nếu listener đã chạy nhưng token/commands đổi so với lần bật trước,
                        # dừng listener cũ để bật lại với cấu hình mới - tránh trường hợp user
                        # sửa token rồi bấm Chạy lại mà listener cũ vẫn dùng token cũ.
                        if is_listener_running(workflow_id):
                            cur_cfg = get_listener_config(workflow_id) or {}
                            if cur_cfg.get("token") != tg_token or cur_cfg.get("commands") != tg_commands:
                                if log_fn:
                                    log_fn(bid, "info", "🔄 Bot Token/lệnh đã đổi - khởi động lại Listener...")
                                _stop_telegram_listener_sync(workflow_id, log_fn=log_fn)
                                # Đợi ngắn để bảng _active_listeners được dọn
                                import time as _t_wait
                                for _ in range(20):
                                    if not is_listener_running(workflow_id):
                                        break
                                    _t_wait.sleep(0.1)

                        if not is_listener_running(workflow_id):
                            def _run_listener_in_thread():
                                import asyncio as _aio
                                from services.telegram_listener import _active_listeners as _listeners_map, _stop_events as _stops_map, _active_configs as _cfg_map
                                loop = _aio.new_event_loop()
                                _aio.set_event_loop(loop)
                                try:
                                    # start_telegram_listener chỉ tạo task nền rồi trả về ngay,
                                    # nên phải tự giữ loop sống bằng cách run_until_complete
                                    # chính task đó - nếu không loop đóng lại và task bị hủy ngay.
                                    loop.run_until_complete(start_telegram_listener(
                                        project_id=project_id,
                                        workflow_id=workflow_id,
                                        workflow_name=workflow_name,
                                        graph_json=graph_json,
                                        bot_token=tg_token,
                                        commands=tg_commands,
                                    ))
                                    task = _listeners_map.get(workflow_id)
                                    if task:
                                        try:
                                            loop.run_until_complete(task)
                                        except _aio.CancelledError:
                                            pass
                                finally:
                                    # Task đã kết thúc (bị dừng/hủy/lỗi) - dọn khỏi bảng theo dõi
                                    # để is_listener_running() phản ánh đúng trạng thái thật.
                                    _listeners_map.pop(workflow_id, None)
                                    _stops_map.pop(workflow_id, None)
                                    _cfg_map.pop(workflow_id, None)
                                    loop.close()

                            _t = _threading_listener.Thread(
                                target=_run_listener_in_thread,
                                daemon=True,
                                name=f"tg-listener-{workflow_id}",
                            )
                            _t.start()

                            _set_workflow_listener_flag(workflow_id, True)

                            if log_fn:
                                log_fn(bid, "success", f"✅ Listener đã được bật. Đang lắng nghe tin nhắn Telegram...")
                        else:
                            _set_workflow_listener_flag(workflow_id, True)
                            if log_fn:
                                log_fn(bid, "info", f"ℹ️ Listener đã chạy sẵn.")
                    except Exception as e:
                        if log_fn:
                            log_fn(bid, "warning", f"⚠ Lỗi khi bật listener: {e}")

                    # Listener chỉ được bật ở đây (từ khối Start) và chỉ tắt khi
                    # người dùng bấm Dừng - workflow giữ trạng thái RUNNING trong
                    # lúc chờ, không tự đánh dấu "hoàn thành" khi chưa chạm khối End.
                    import time as _time_listener
                    while True:
                        if stop_event and stop_event.is_set():
                            if log_fn:
                                log_fn(bid, "warning", f"⏹ Đang tắt Listener theo yêu cầu người dùng...")
                            _stop_telegram_listener_sync(workflow_id, log_fn=log_fn)
                            _set_workflow_listener_flag(workflow_id, False)
                            final_status = "stopped"
                            break
                        _time_listener.sleep(0.5)
                    break
            elif btype == "telegram":
                bot_token = bdata.get("telegramBotToken", "").strip()
                chat_id = bdata.get("telegramChatId", "").strip()
                text = bdata.get("telegramMessage", "")
                parse_mode = bdata.get("telegramParseMode", "")
                telegram_attachments = bdata.get("telegramAttachments", [])
                telegram_action = bdata.get("telegramAction", "send")
                msg_id = str(bdata.get("telegramMessageId", ""))
                
                if not bot_token or not chat_id:
                    if log_fn:
                        log_fn(bid, "error", f"❌ [Telegram] {label} - Thiếu Bot Token hoặc Chat ID")
                    final_status = "error"
                    break

                # --- Resolve attachment file paths ---
                resolved_files = []
                for att in telegram_attachments:
                    att_name = att
                    if not att_name:
                        continue
                    # Tìm trong output trước, rồi input
                    att_path = wf_dir / "output" / att_name
                    if not att_path.exists():
                        att_path = input_dir / att_name
                    if att_path.exists() and att_path.is_file():
                        resolved_files.append((att_name, att_path))
                    else:
                        if log_fn:
                            log_fn(bid, "warning", f"⚠ [Telegram] Không tìm thấy file đính kèm: {att_name}")

                def _tg_send_message(bot_token, chat_id, text, parse_mode, reply_to=None):
                    """Gửi tin nhắn văn bản qua Telegram sendMessage"""
                    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                    payload = {"chat_id": chat_id, "text": text}
                    if parse_mode:
                        payload["parse_mode"] = parse_mode
                    if reply_to:
                        payload["reply_to_message_id"] = reply_to
                    req_data = json.dumps(payload).encode('utf-8')
                    req = urllib.request.Request(url, data=req_data, headers={'Content-Type': 'application/json'})
                    with urllib.request.urlopen(req, timeout=15) as response:
                        return json.loads(response.read().decode('utf-8'))
                        
                def _tg_edit_message_text(bot_token, chat_id, text, message_id, parse_mode):
                    """Sửa tin nhắn văn bản qua Telegram editMessageText"""
                    url = f"https://api.telegram.org/bot{bot_token}/editMessageText"
                    payload = {"chat_id": chat_id, "message_id": message_id, "text": text}
                    if parse_mode:
                        payload["parse_mode"] = parse_mode
                    req_data = json.dumps(payload).encode('utf-8')
                    req = urllib.request.Request(url, data=req_data, headers={'Content-Type': 'application/json'})
                    with urllib.request.urlopen(req, timeout=15) as response:
                        return json.loads(response.read().decode('utf-8'))

                def _tg_send_document(bot_token, chat_id, file_path, file_name, caption="", parse_mode="", reply_to=None):
                    """Gửi file qua Telegram sendDocument (multipart/form-data)"""
                    import uuid
                    boundary = uuid.uuid4().hex
                    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
                    
                    lines = []
                    # chat_id field
                    lines.append(f"--{boundary}".encode())
                    lines.append(b'Content-Disposition: form-data; name="chat_id"')
                    lines.append(b'')
                    lines.append(chat_id.encode())
                    # document field
                    lines.append(f"--{boundary}".encode())
                    lines.append(f'Content-Disposition: form-data; name="document"; filename="{file_name}"'.encode())
                    lines.append(b'Content-Type: application/octet-stream')
                    lines.append(b'')
                    with open(file_path, 'rb') as fp:
                        file_data = fp.read()
                    lines.append(file_data)
                    # caption field (optional)
                    if caption:
                        lines.append(f"--{boundary}".encode())
                        lines.append(b'Content-Disposition: form-data; name="caption"')
                        lines.append(b'')
                        lines.append(caption.encode())
                    # parse_mode field (optional)
                    if parse_mode:
                        lines.append(f"--{boundary}".encode())
                        lines.append(b'Content-Disposition: form-data; name="parse_mode"')
                        lines.append(b'')
                        lines.append(parse_mode.encode())
                    # reply_to field (optional)
                    if reply_to:
                        lines.append(f"--{boundary}".encode())
                        lines.append(b'Content-Disposition: form-data; name="reply_to_message_id"')
                        lines.append(b'')
                        lines.append(str(reply_to).encode())
                    lines.append(f"--{boundary}--".encode())
                    
                    body = b"\r\n".join(lines)
                    req = urllib.request.Request(url, data=body, headers={
                        'Content-Type': f'multipart/form-data; boundary={boundary}'
                    })
                    with urllib.request.urlopen(req, timeout=60) as response:
                        return json.loads(response.read().decode('utf-8'))
                    
                if log_fn:
                    log_fn(bid, "info", f"✉️ [Telegram] {label} - Đang gửi tin nhắn tới {chat_id}...")
                
                try:
                    telegram_error = False
                    last_message_id = None
                    
                    if telegram_action == "edit":
                        if not msg_id:
                            if log_fn:
                                log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi: Chế độ 'Sửa tin nhắn' yêu cầu Message ID hợp lệ.")
                            final_status = "error"
                            break
                        try:
                            msg_id_int = int(msg_id)
                        except ValueError:
                            if log_fn: log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi: Message ID phải là số (hiện tại là '{msg_id}'). Hãy kiểm tra lại biến.")
                            final_status = "error"
                            break
                        res = _tg_edit_message_text(bot_token, chat_id, text, msg_id_int, parse_mode)
                        if not res.get("ok"):
                            if log_fn: log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi sửa tin nhắn: {res.get('description')}")
                            telegram_error = True
                        else:
                            last_message_id = res.get("result", {}).get("message_id")
                    else:
                        reply_to_id = None
                        if telegram_action == "reply" and msg_id:
                            try:
                                reply_to_id = int(msg_id)
                            except ValueError:
                                if log_fn: log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi: Message ID phải là số (hiện tại là '{msg_id}').")
                                final_status = "error"
                                break
                        
                        if not resolved_files:
                            # Không có file đính kèm
                            res = _tg_send_message(bot_token, chat_id, text, parse_mode, reply_to_id)
                            if not res.get("ok"):
                                if log_fn:
                                    log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi API: {res.get('description')}")
                                telegram_error = True
                            else:
                                last_message_id = res.get("result", {}).get("message_id")
                        elif len(resolved_files) == 1:
                            # 1 file
                            fname, fpath = resolved_files[0]
                            if log_fn:
                                log_fn(bid, "info", f"📎 [Telegram] {label} - Đính kèm file: {fname}")
                            res = _tg_send_document(bot_token, chat_id, str(fpath), fname, caption=text, parse_mode=parse_mode, reply_to=reply_to_id)
                            if not res.get("ok"):
                                if log_fn:
                                    log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi gửi file {fname}: {res.get('description')}")
                                telegram_error = True
                            else:
                                last_message_id = res.get("result", {}).get("message_id")
                        else:
                            # Nhiều file
                            if text.strip():
                                res = _tg_send_message(bot_token, chat_id, text, parse_mode, reply_to_id)
                                if not res.get("ok"):
                                    if log_fn:
                                        log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi gửi tin nhắn: {res.get('description')}")
                                    telegram_error = True
                                else:
                                    last_message_id = res.get("result", {}).get("message_id")
                            
                            if not telegram_error:
                                for i, (fname, fpath) in enumerate(resolved_files):
                                    if log_fn:
                                        log_fn(bid, "info", f"📎 [Telegram] {label} - Đính kèm file: {fname}")
                                    # File đầu tiên reply to the message (nếu text trống và có reply_to_id)
                                    cur_reply_to = reply_to_id if (i == 0 and not text.strip()) else None
                                    res = _tg_send_document(bot_token, chat_id, str(fpath), fname, reply_to=cur_reply_to)
                                    if not res.get("ok"):
                                        if log_fn:
                                            log_fn(bid, "error", f"❌ [Telegram] {label} - Lỗi gửi file {fname}: {res.get('description')}")
                                        telegram_error = True
                                        break
                                    if not last_message_id:
                                        last_message_id = res.get("result", {}).get("message_id")
                    
                    if telegram_error:
                        final_status = "error"
                        break
                    else:
                        if log_fn:
                            log_fn(bid, "success", f"✅ [Telegram] {label} - Đã gửi thánh công! (message_id={last_message_id})")
                        if isinstance(current_input, dict):
                            # Giữ lại message_id cũ (tin nhắn người dùng), thêm sent_message_id (tin vừa gửi ra)
                            current_input["sent_message_id"] = last_message_id
                            current_input["chat_id"] = chat_id
                        else:
                            current_input = {"sent_message_id": last_message_id, "message_id": last_message_id, "chat_id": chat_id}

                        sent_id_var = bdata.get("telegramSentMessageIdVarName", "").strip()
                        if sent_id_var:
                            workflow_env[sent_id_var] = current_input.get("sent_message_id")
                        chat_id_var = bdata.get("telegramChatIdVarName", "").strip()
                        if chat_id_var:
                            workflow_env[chat_id_var] = current_input.get("chat_id")
                except Exception as e:
                    if log_fn:
                        log_fn(bid, "error", f"❌ [Telegram] {label} - Gửi thất bại: {str(e)}")
                    final_status = "error"
                    break
            elif btype == "email":
                mail_host = bdata.get("mailHost", "").strip()
                mail_port = int(bdata.get("mailPort", 465) or 465)
                mail_user = bdata.get("mailUser", "").strip()
                mail_pass = bdata.get("mailPass", "").replace(" ", "")
                mail_to = bdata.get("mailTo", "").strip()
                mail_cc = bdata.get("mailCc", "").strip()
                mail_subject = bdata.get("mailSubject", "").strip()
                mail_body = bdata.get("mailBody", "")
                mail_attachments = bdata.get("mailAttachments", [])

                final_to = mail_to
                final_cc = mail_cc
                final_subject = mail_subject
                final_body = mail_body

                if log_fn:
                    log_fn(bid, "info", f"📧 [Email] {label} - Đang gửi thư tới {final_to}...")

                msg = EmailMessage()
                msg['Subject'] = final_subject
                msg['From'] = mail_user
                msg['To'] = final_to
                if final_cc:
                    msg['Cc'] = final_cc
                
                # Check if body contains HTML tags
                if "<" in final_body and ">" in final_body:
                    msg.set_content(final_body, subtype='html')
                else:
                    msg.set_content(final_body)

                # Attachments
                for att in mail_attachments:
                    att_name = att
                    if not att_name: continue
                    # try INPUT_DIR then OUTPUT_DIR
                    att_path = input_dir / att_name
                    if not att_path.exists():
                        att_path = wf_dir / "output" / att_name
                    
                    if att_path.exists() and att_path.is_file():
                        import mimetypes
                        ctype, encoding = mimetypes.guess_type(str(att_path))
                        if ctype is None or encoding is not None:
                            ctype = 'application/octet-stream'
                        maintype, subtype = ctype.split('/', 1)
                        with open(att_path, 'rb') as fp:
                            msg.add_attachment(fp.read(), maintype=maintype, subtype=subtype, filename=att_name)
                    else:
                        if log_fn:
                            log_fn(bid, "warning", f"⚠ [Email] Không tìm thấy file đính kèm: {att_name}")

                try:
                    # Decide SSL or TLS based on port
                    if mail_port == 465:
                        smtp = smtplib.SMTP_SSL(mail_host, mail_port, timeout=15)
                    else:
                        smtp = smtplib.SMTP(mail_host, mail_port, timeout=15)
                        smtp.starttls()
                    
                    smtp.login(mail_user, mail_pass)
                    smtp.send_message(msg)
                    smtp.quit()

                    if log_fn:
                        log_fn(bid, "success", f"✅ [Email] {label} - Đã gửi thư thánh công!")
                except Exception as e:
                    if log_fn:
                        log_fn(bid, "error", f"❌ [Email] {label} - Lỗi gửi thư: {str(e)}")
                    final_status = "error"
                    break
            elif btype == "delete_files":
                import shutil
                delete_input = bdata.get("delete_input", False)
                delete_output = bdata.get("delete_output", False)

                if delete_input:
                    try:
                        if input_dir.exists():
                            for item in input_dir.iterdir():
                                if item.name == "input.json":
                                    continue
                                if item.is_file():
                                    item.unlink()
                                elif item.is_dir():
                                    shutil.rmtree(item)
                        if log_fn:
                            log_fn(bid, "success", f"✅ [Xóa] Đã dọn dẹp thư mục Input.")
                    except Exception as e:
                        if log_fn:
                            log_fn(bid, "error", f"❌ [Xóa] Lỗi xóa Input: {e}")
                
                if delete_output:
                    try:
                        out_dir = wf_dir / "output"
                        if out_dir.exists():
                            for item in out_dir.iterdir():
                                if item.is_file():
                                    item.unlink()
                                elif item.is_dir():
                                    shutil.rmtree(item)
                        if log_fn:
                            log_fn(bid, "success", f"✅ [Xóa] Đã dọn dẹp thư mục Output.")
                    except Exception as e:
                        if log_fn:
                            log_fn(bid, "error", f"❌ [Xóa] Lỗi xóa Output: {e}")
                            
            elif btype == "browser":
                steps = bdata.get("steps", [])
                headless = not bdata.get("debugMode", False)
                if not steps:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có bước nào, bỏ qua")
                    continue_branch = False
                else:
                    if log_fn:
                        log_fn(bid, "info", f"🌐 Đang chạy Browser: {label}...")
                    
                    from services.browser_executor import run_browser_block
                    
                    start_b = datetime.now()
                    try:
                        output_dir = wf_dir / "output"
                        output_dir.mkdir(exist_ok=True)
                        
                        # Không nội suy trước ở đây nữa vì sẽ gây lỗi lấy nhầm dữ liệu
                        # vòng lặp cũ (Race condition với các biến nội bộ sinh ra trong cùng block).
                        # Truyền nguyên steps và gom biến toàn cục vào cho browser_executor tự xử.
                        merged_input = dict(workflow_env)
                        if isinstance(current_input, dict):
                            merged_input.update(current_input)
                            
                        b_result = run_browser_block(
                            block_id=bid,
                            workflow_id=workflow_id,
                            run_id=run_id,
                            steps=steps,
                            input_data=merged_input,
                            headless=headless,
                            log_callback=log_fn,
                            output_dir=str(output_dir).replace('\\', '/'),
                            stop_event=stop_event,
                            browser_profile_dir=str(browser_profile_dir).replace('\\', '/'),
                        )
                        if not b_result.get("success"):
                            if b_result.get("stopped"):
                                # Không finish_run trực tiếp - để rơi xuống cuối vòng lặp,
                                # nơi đã có sẵn log "⏹ Đã dừng sau Xms" cho final_status=stopped.
                                final_status = "stopped"
                                break
                            if log_fn:
                                log_fn("system", "error", f"❌ Workflow thất bại (Browser lỗi)")
                            if handle_workflow_error(b_result.get("error"), bid, label):
                                return
                            else:
                                continue
                        else:
                            if b_result.get("output_data") is not None:
                                browser_out = b_result["output_data"]
                                # Merge: giữ lại tất cả biến cũ (chat_id, message_id, sent_message_id...)
                                # rồi cập nhật thêm dữ liệu mới từ Browser vào
                                if isinstance(current_input, dict) and isinstance(browser_out, dict):
                                    merged = dict(current_input)
                                    merged.update(browser_out)
                                    current_input = merged
                                else:
                                    current_input = browser_out
                    except Exception as e:
                        if log_fn:
                            log_fn("system", "error", f"❌ Lỗi ngoại lệ Browser: {e}")
                        if handle_workflow_error(str(e), bid, label):
                            return
                        else:
                            continue
            elif btype == "python":
                code = bdata.get("code", "").strip()
                if not code:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có code")
                    continue_branch = False
                else:
                    if log_fn:
                        log_fn(bid, "info", f"⚡ Đang chạy: {label}...")
                    success, output, error, duration = run_python_block_sync(
                        project_id, bid, workflow_id, code, current_input, 
                        timeout=7200, label=label, log_fn=log_fn, input_dir=str(input_dir),
                        stop_event=stop_event
                    )
                    if not success:
                        if error == "stopped":
                            final_status = "stopped"
                            break
                        if log_fn:
                            log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                        if handle_workflow_error(error, bid, label):
                            return
                        else:
                            continue
                    current_input = output
            elif btype == "sql_to_excel":
                sql_query = bdata.get("sqlQuery", "").strip()
                excel_filename = bdata.get("excelFileName", "sqltoexcel.xlsx").strip()
                saved_connection_id = bdata.get("sqlToExcelSavedConnectionId", "").strip()
                if not excel_filename:
                    excel_filename = "sqltoexcel.xlsx"

                db_config = get_saved_db_connection(saved_connection_id)

                if not sql_query:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có câu lệnh SQL")
                    continue_branch = False
                elif not db_config:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Block [{label}] chưa chọn Kết nối Database")
                    if handle_workflow_error("Chưa chọn Kết nối Database cho khối SQL to Excel", bid, label):
                        return
                    else:
                        continue
                else:
                    conn_str = build_conn_str(db_config)
                    db_type = db_config.get("db_type")
                    packages_to_install = ["pandas", "sqlalchemy", "openpyxl"]
                    if db_type == "postgresql": packages_to_install.append("psycopg2-binary")
                    elif db_type == "mysql": packages_to_install.extend(["pymysql", "cryptography"])
                    else: packages_to_install.append("pyodbc")

                    ensure_packages(project_id, packages_to_install, log_fn, bid, label, stop_event)

                    if log_fn:
                        log_fn(bid, "info", f"⚡ Đang chạy SQL to Excel: {label}...")

                    code = f'''
import pandas as pd
import sqlalchemy
import os

conn_str = {conn_str!r}

print("Đang kết nối CSDL và thực thi câu lệnh SQL...")
engine = sqlalchemy.create_engine(conn_str)
df = pd.read_sql("""{sql_query}""", engine)

out_file = {excel_filename!r}
out_path = os.path.join(OUTPUT_DIR, out_file)

print(f"Đã tải {{len(df)}} dòng dữ liệu! Đang lưu vào {{out_path}}...")
df.to_excel(out_path, index=False)

output_data = {{"file_name": out_file}}
'''
                    success, output, error, duration = run_python_block_sync(
                        project_id, bid, workflow_id, code, current_input,
                        timeout=7200, label=label, log_fn=log_fn, input_dir=str(input_dir),
                        stop_event=stop_event
                    )
                    if not success:
                        if log_fn:
                            log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                        if handle_workflow_error(error, bid, label):
                            return
                        else:
                            continue
                    current_input = output
            elif btype == "merge_excel":
                header_rows = int(bdata.get("headerRows", 3))
                excel_filename = bdata.get("excelFileName", "merged_excel.xlsx").strip()
                merge_mode = bdata.get("mergeMode") or ("all_input" if bdata.get("mergeAllInput", True) else "custom")
                merge_all_input = merge_mode in ("all_input", "all_output")
                selected_files = bdata.get("selectedFiles", [])
                if not excel_filename:
                    excel_filename = "merged_excel.xlsx"
                
                # Nếu bật chọn tất cả Input: tự động quét toàn bộ file trong INPUT_DIR
                if merge_all_input:
                    selected_files = None  # Sẽ tự quét trong code Python
                elif not selected_files:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Merge Excel: Bạn chưa chọn file nào để gộp!")
                    if handle_workflow_error("No files selected", bid, label):
                        return
                    else:
                        continue
                
                packages_to_install = ["pandas", "openpyxl", "xlrd"]
                ensure_packages(project_id, packages_to_install, log_fn, bid, label, stop_event)
                
                if log_fn:
                    log_fn(bid, "info", f"⚡ Đang chạy Merge Excel: {label}...")
                
                if merge_all_input:
                    file_list_code = """
# T\u1ef1 đ\u1ed9ng qu\u00e9t t\u1ea5t c\u1ea3 file .xlsx/.csv trong INPUT_DIR
file_list = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx') or f.endswith('.csv')])
print(f"T\u1ef1 đ\u1ed9ng ph\u00e1t hi\u1ec7n {len(file_list)} file trong th\u01b0 m\u1ee5c Input.")
"""
                else:
                    _fl = selected_files
                    file_list_code = f"""file_list = {_fl!r}\nprint(f\"\u0110ang gh\u00e9p {{len(file_list)}} file \u0111\u00e3 ch\u1ecdn.\")"""
                if merge_mode == "all_output":
                    file_list_code = file_list_code.replace("INPUT_DIR", "OUTPUT_DIR").replace("Input.", "Output.")

                code = f'''
import os
import warnings
import pandas as pd
import openpyxl

# Tắt cảnh báo openpyxl "Workbook contains no default style" (vô hại, chỉ gây nhiễu log)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

{file_list_code}

if not file_list:
    raise ValueError(f"Không tìm thấy file .xlsx nào trong Dữ liệu Workflow (INPUT_DIR)")

print(f"Bắt đầu ghép. Giữ nguyên {{ {header_rows} }} dòng tiêu đề từ file đầu tiên...")

file1_path = os.path.join(INPUT_DIR, file_list[0])
if not os.path.exists(file1_path):
    file1_path = os.path.join(OUTPUT_DIR, file_list[0])
    
if not os.path.exists(file1_path):
    raise FileNotFoundError(f"Không tìm thấy file Gốc: {{file_list[0]}}")

# 1. Dùng openpyxl mở File 1 để giữ trọn vẹn Format tiêu đề
print(f"Đang xử lý định dạng tiêu đề từ File 1: {{file_list[0]}}")
wb = openpyxl.load_workbook(file1_path)
ws = wb.active

# Xóa toàn bộ dữ liệu bên dưới dòng tiêu đề (Để lấy format chuẩn cho phần data sau này)
if ws.max_row > {header_rows}:
    ws.delete_rows({header_rows} + 1, ws.max_row)

# 2. Dùng pandas đọc phần Data của TẤT CẢ các file (Bao gồm cả file 1)
dfs = []
for filename in file_list:
    file_path = os.path.join(INPUT_DIR, filename)
    if not os.path.exists(file_path):
        file_path = os.path.join(OUTPUT_DIR, filename)
        
    print(f"   + Đang đọc dữ liệu: {{filename}}")
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Không tìm thấy file: {{filename}}")
        
    # Bỏ qua header_rows, chỉ lấy Data thô
    df = pd.read_excel(file_path, header=None, skiprows={header_rows})
    dfs.append(df)

print("Đang gộp và đắp dữ liệu vào bảng...")
merged_df = pd.concat(dfs, ignore_index=True)

# 3. Ghi dữ liệu vào Workbook đã có sẵn tiêu đề đẹp
# Chuyển DataFrame thành list các dòng để append vào openpyxl siêu nhanh
for row in merged_df.itertuples(index=False, name=None):
    ws.append(row)

out_file = {excel_filename!r}
out_path = os.path.join(OUTPUT_DIR, out_file)

print(f"Đang lưu file kết quả (Gồm {header_rows} dòng tiêu đề + {{len(merged_df)}} dòng dữ liệu) vào {{out_path}}...")
try:
    wb.save(out_path)
except PermissionError:
    raise PermissionError(f"Không thể lưu đè file {{out_file}} do file này đang được mở trong Excel. Vui lòng đóng file Excel và chạy lại.")
output_data = {{"file_name": out_file}}
'''
                success, output, error, duration = run_python_block_sync(
                    project_id, bid, workflow_id, code, current_input,
                    timeout=7200, label=label, log_fn=log_fn, input_dir=str(input_dir),
                    stop_event=stop_event
                )
                if not success:
                    if log_fn:
                        log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                    if handle_workflow_error(error, bid, label):
                        return
                    else:
                        continue
                current_input = output
            elif btype == "pivot_excel":
                excel_filename = bdata.get("excelFileName", "pivot.xlsx").strip()
                selected_files = bdata.get("pivotInputFiles", [])
                pivot_index = bdata.get("pivotIndex", "")
                pivot_columns = bdata.get("pivotColumns", "")
                pivot_values = bdata.get("pivotValues", "")
                pivot_agg = bdata.get("pivotAgg", "sum")
                pivot_fillna = bdata.get("pivotFillNa", True)
                pivot_grand_total = bdata.get("pivotGrandTotal", True)
                pivot_header_row = int(bdata.get("pivotHeaderRow", 1))
                pivot_enable_sort = bdata.get("pivotEnableSort", False)
                pivot_sort_column = bdata.get("pivotSortColumn", "").strip()
                pivot_sort_order = bdata.get("pivotSortOrder", "asc")
                pivot_sort_custom = bdata.get("pivotSortCustom", [])
                
                if not excel_filename:
                    excel_filename = "pivot.xlsx"
                
                if not selected_files:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Pivot Excel: Bạn chưa chọn file nào để tổng hợp!")
                    if handle_workflow_error("No files selected", bid, label):
                        return
                    else:
                        continue
                
                packages_to_install = ["pandas", "openpyxl", "xlrd"]
                ensure_packages(project_id, packages_to_install, log_fn, bid, label, stop_event)
                
                if log_fn:
                    log_fn(bid, "info", f"📊 Đang chạy Pivot Excel: {label}...")
                
                code = f'''
import os
import warnings
import pandas as pd

# Tắt cảnh báo openpyxl "Workbook contains no default style" (vô hại, chỉ gây nhiễu log)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

print(f"Đang tìm kiếm {{len({selected_files!r})}} file Excel trong thư mục: {{INPUT_DIR}}...")
file_list = {selected_files!r}

if not file_list:
    raise ValueError(f"Không tìm thấy file nào để xử lý.")

dfs = []
for filename in file_list:
    file_path = os.path.join(INPUT_DIR, filename)
    if not os.path.exists(file_path):
        file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Không tìm thấy file: {{filename}}")
    
    print(f"Đang đọc dữ liệu: {{filename}} (header dòng {{ {pivot_header_row} }})")
    pd_header_idx = max(0, {pivot_header_row} - 1)
    if filename.endswith(".csv"):
        df = pd.read_csv(file_path, header=pd_header_idx)
    else:
        df = pd.read_excel(file_path, header=pd_header_idx)
    dfs.append(df)

df = pd.concat(dfs, ignore_index=True)

def col2num(col_str):
    num = 0
    for c in str(col_str).upper():
        if 'A' <= c <= 'Z':
            num = num * 26 + (ord(c) - ord('A') + 1)
        else:
            return -1
    return num - 1

# Lọc và dọn dẹp biến đầu vào
def parse_cols(col_data):
    if not col_data: return []
    if isinstance(col_data, str):
        cols = [c.strip() for c in col_data.split(',')]
    elif isinstance(col_data, list):
        cols = [str(c).strip() for c in col_data]
    else:
        return []
        
    parsed = []
    for c in cols:
        if not c: continue
        if c in df.columns:
            parsed.append(c)
        else:
            idx = col2num(c)
            if 0 <= idx < len(df.columns):
                parsed.append(df.columns[idx])
            else:
                print(f"Cảnh báo: Không tìm thấy cột '{{c}}'")
    return parsed

idx_cols = parse_cols({pivot_index!r})
col_cols = parse_cols({pivot_columns!r})
val_cols = parse_cols({pivot_values!r})

if not val_cols:
    print("Cảnh báo: Không có cột Giá trị (Values) nào hợp lệ. Sẽ dùng toàn bộ cột số.")
    val_cols = df.select_dtypes(include='number').columns.tolist()
if not idx_cols and not col_cols:
    raise ValueError("Phải nhập ít nhất 1 trường Dòng (Rows) hoặc Cột (Columns).")

sort_col_parsed = parse_cols({pivot_sort_column!r})
sort_col = sort_col_parsed[0] if sort_col_parsed else None

if {pivot_enable_sort} and sort_col and sort_col in df.columns:
    print(f"Đang phân loại và sắp xếp cột '{{sort_col}}' theo chiều: {{ {pivot_sort_order!r} }}")
    if {pivot_sort_order!r} == 'custom':
        custom_order = {pivot_sort_custom!r}
        # Xử lý Pandas4Warning: bổ sung các giá trị còn thiếu vào cuối danh sách categories
        unique_vals = df[sort_col].dropna().unique().tolist()
        # Ép kiểu custom_order về cùng kiểu với dữ liệu để so sánh chính xác nếu cần
        try:
            custom_order = pd.Series(custom_order).astype(df[sort_col].dtype).tolist()
        except:
            pass
        missing_cats = [x for x in unique_vals if x not in custom_order]
        custom_order.extend(missing_cats)
        df[sort_col] = pd.Categorical(df[sort_col], categories=custom_order, ordered=True)
    elif {pivot_sort_order!r} == 'asc':
        cats = sorted(df[sort_col].dropna().unique().tolist(), key=str)
        df[sort_col] = pd.Categorical(df[sort_col], categories=cats, ordered=True)
    elif {pivot_sort_order!r} == 'desc':
        cats = sorted(df[sort_col].dropna().unique().tolist(), key=str, reverse=True)
        df[sort_col] = pd.Categorical(df[sort_col], categories=cats, ordered=True)

print(f"Đang Pivot với Dòng={{idx_cols}}, Cột={{col_cols}}, Giá trị={{val_cols}}, Hàm={pivot_agg!r}")

pivot_df = pd.pivot_table(
    df, 
    values=val_cols, 
    index=idx_cols if idx_cols else None, 
    columns=col_cols if col_cols else None, 
    aggfunc={pivot_agg!r}, 
    margins={pivot_grand_total},
    margins_name="Tổng Cộng"
)

if {pivot_fillna}:
    pivot_df = pivot_df.fillna(0)

# Làm phẳng cột đa cấp (nếu có)
if isinstance(pivot_df.columns, pd.MultiIndex):
    new_cols = []
    for col in pivot_df.columns.values:
        valid_parts = [str(c) for c in col if str(c) != '']
        if not valid_parts:
            new_cols.append('')
        elif len(valid_parts) > 1 and valid_parts[0] in val_cols:
            new_cols.append('_'.join(valid_parts[1:]))
        else:
            new_cols.append('_'.join(valid_parts))
    pivot_df.columns = new_cols

out_file = {excel_filename!r}
out_path = os.path.join(OUTPUT_DIR, out_file)

# Reset index để lưu ra file excel đẹp hơn
pivot_df.reset_index(inplace=True)

print(f"Đang lưu file Pivot kết quả vào {{out_path}}...")
try:
    pivot_df.to_excel(out_path, index=False)
except PermissionError:
    raise PermissionError(f"Không thể lưu đè file {{out_file}} do file này đang được mở trong Excel.")

output_data = {{"file_name": out_file}}
'''
                success, output, error, duration = run_python_block_sync(
                    project_id, bid, workflow_id, code, current_input,
                    timeout=7200, label=label, log_fn=log_fn, input_dir=str(input_dir),
                    stop_event=stop_event
                )
                if not success:
                    if log_fn:
                        log_fn("system", "error", f"❌ Workflow thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                    if handle_workflow_error(error, bid, label):
                        return
                    else:
                        continue
                current_input = output
            elif btype == "excel_to_sql":
                input_file = bdata.get("excelToSqlInputFile", "").strip()
                saved_connection_id = bdata.get("excelToSqlSavedConnectionId", "").strip()
                table_name = bdata.get("excelToSqlTableName", "").strip()
                import_mode = bdata.get("excelToSqlImportMode", "append")
                mapping = bdata.get("excelToSqlMapping", {})

                header_row_str = str(bdata.get("excelToSqlHeaderRow", "1"))
                h_parts = [p.strip() for p in header_row_str.replace('-', ',').split(',')]
                h_indices = []
                for p in h_parts:
                    if p.isdigit():
                        h_indices.append(max(0, int(p) - 1))
                if not h_indices:
                    header_row_val = 0
                elif len(h_indices) == 1:
                    header_row_val = h_indices[0]
                else:
                    header_row_val = h_indices

                db_config = get_saved_db_connection(saved_connection_id)

                if not input_file or not table_name:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Thiếu cấu hình: File nguồn hoặc Bảng đích")
                    if handle_workflow_error("Chưa cấu hình đủ bảng đích hoặc file nguồn", bid, label):
                        return
                    else:
                        continue
                elif not db_config:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Block [{label}] chưa chọn Kết nối Database")
                    if handle_workflow_error("Chưa chọn Kết nối Database cho khối Excel to SQL", bid, label):
                        return
                    else:
                        continue

                conn_str = build_conn_str(db_config)
                db_type = db_config.get("db_type")
                packages_to_install = ["pandas", "openpyxl", "sqlalchemy", "xlrd"]
                if db_type == "postgresql": packages_to_install.append("psycopg2-binary")
                elif db_type == "mysql": packages_to_install.extend(["pymysql", "cryptography"])
                else: packages_to_install.append("pyodbc")
                ensure_packages(project_id, packages_to_install, log_fn, bid, label, stop_event)

                if log_fn:
                    log_fn(bid, "info", "⚡ Đang Import Excel vào SQL Server: " + label + "...")

                code = f'''
import os
import warnings
import pandas as pd
from sqlalchemy import create_engine, text

# Suppress openpyxl style warnings (file không có default style vẫn đọc được)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

input_file = {input_file!r}
file_path = os.path.join(INPUT_DIR, input_file)
if not os.path.exists(file_path):
    file_path = os.path.join(OUTPUT_DIR, input_file)

if not os.path.exists(file_path):
    raise FileNotFoundError(f"Không tìm thấy file Excel: {{input_file}}")

print(f"Đang đọc file Excel (dòng tiêu đề = {header_row_str} theo Excel, pandas header={header_row_val})...")
if file_path.endswith(".csv"):
    df = pd.read_csv(file_path, header={header_row_val})
else:
    df = pd.read_excel(file_path, header={header_row_val})

if isinstance(df.columns, pd.MultiIndex):
    flat_cols = []
    for col in df.columns:
        clean_levels = [str(c).strip() for c in col if not str(c).startswith('Unnamed:')]
        flat_cols.append('_'.join(clean_levels) if clean_levels else 'Unnamed')
    df.columns = flat_cols

conn_str = {conn_str!r}
engine = create_engine(conn_str, fast_executemany=True)

mapping = {mapping!r}
print(f"File Excel có {{len(df)}} dòng dữ liệu, {{len(df.columns)}} cột.")
print(f"10 cột đầu của Excel: {{list(df.columns[:10])}}")

if mapping:
    print(f"Đang chuẩn hóa dữ liệu theo Mapping: {{len(mapping)}} cột SQL...")
    # Khởi tạo với đúng số dòng từ df (tránh lỗi 0 dòng khi cột không khớp)
    sql_df = pd.DataFrame(index=range(len(df)))
    matched_count = 0
    unmatched = []
    for sql_col, excel_col in mapping.items():
        if excel_col and excel_col in df.columns:
            sql_df[sql_col] = df[excel_col].values
            matched_count += 1
        elif excel_col:
            sql_df[sql_col] = None
            unmatched.append(f"'{{excel_col}}' -> '{{sql_col}}'")
        else:
            sql_df[sql_col] = None
    print(f"Đã khớp {{matched_count}}/{{len(mapping)}} cột. Tổng dòng cần import: {{len(sql_df)}}")
    if unmatched:
        print(f"⚠ {{len(unmatched)}} cột Excel không khớp (sẽ để NULL): {{unmatched[:5]}}")
else:
    print("Không có Mapping cụ thể, đẩy toàn bộ dữ liệu Excel hiện có...")
    sql_df = df.copy()

table_name = {table_name!r}
import_mode = {import_mode!r}

from sqlalchemy import inspect
insp = inspect(engine)
try:
    columns = insp.get_columns(table_name)
    date_cols = [c['name'] for c in columns if 'DATE' in str(c['type']).upper() or 'TIME' in str(c['type']).upper()]
    
    if date_cols:
        print(f"Phát hiện {{len(date_cols)}} cột thời gian trong SQL. Đang ép kiểu nghiêm ngặt (dayfirst=True)...")
        for sql_col in date_cols:
            if sql_col in sql_df.columns:
                # errors='raise' để báo lỗi dừng khối ngay lập tức nếu format rác/sai
                sql_df[sql_col] = pd.to_datetime(sql_df[sql_col], dayfirst=True, errors='raise')
except Exception as e:
    raise ValueError(f"Lỗi dữ liệu ngày tháng không hợp lệ (sai định dạng hoặc chứa chữ rác): {{str(e)}}")

# Chuẩn hoá NaN -> None (cho cả các ô trống do pd.to_datetime tạo ra nếu có NaT)
sql_df = sql_df.where(pd.notnull(sql_df), None)

with engine.begin() as connection:
    if import_mode == 'truncate':
        print(f"Đang xoá trắng dữ liệu cũ trong bảng {{table_name}}...")
        connection.execute(text(f"TRUNCATE TABLE [{{table_name}}]"))

print(f"Đang Bulk Insert {{len(sql_df)}} dòng vào bảng {{table_name}}...")
sql_df.to_sql(name=table_name, con=engine, if_exists='append', index=False, chunksize=10000)

print(f"✅ Đã Import thành công {{len(sql_df)}} dòng.")
output_data = {{"rows_inserted": len(sql_df), "table": table_name}}
'''
                success, output, error, duration = run_python_block_sync(
                    project_id, bid, workflow_id, code, current_input,
                    timeout=7200, label=label, log_fn=log_fn, input_dir=str(input_dir),
                    stop_event=stop_event
                )
                if not success:
                    if log_fn:
                        log_fn("system", "error", f"❌ Import thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                    if handle_workflow_error(error, bid, label):
                        return
                    else:
                        continue
                else:
                    current_input = output
                    if isinstance(current_input, dict):
                        rows_var = bdata.get("excelToSqlRowsVarName", "").strip()
                        if rows_var:
                            workflow_env[rows_var] = current_input.get("rows_inserted")
                        table_var = bdata.get("excelToSqlTableVarName", "").strip()
                        if table_var:
                            workflow_env[table_var] = current_input.get("table")
            elif btype == "run_sql_exec":
                sql_command = bdata.get("sqlCommand", "").strip()
                saved_connection_id = bdata.get("sqlExecSavedConnectionId", "").strip()

                db_config = get_saved_db_connection(saved_connection_id)

                if not sql_command:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có câu lệnh SQL/EXEC")
                    continue_branch = False
                elif not db_config:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Block [{label}] chưa chọn Kết nối Database")
                    if handle_workflow_error("Chưa chọn Kết nối Database cho khối Chạy Hàm SQL (EXEC)", bid, label):
                        return
                    else:
                        continue
                else:
                    conn_str = build_conn_str(db_config)
                    db_type = db_config.get("db_type")
                    packages_to_install = ["sqlalchemy"]
                    if db_type == "postgresql": packages_to_install.append("psycopg2-binary")
                    elif db_type == "mysql": packages_to_install.extend(["pymysql", "cryptography"])
                    else: packages_to_install.append("pyodbc")

                    ensure_packages(project_id, packages_to_install, log_fn, bid, label, stop_event)

                    if log_fn:
                        log_fn(bid, "info", "⚡ Đang chạy Hàm/Thủ tục SQL: " + label + "...")

                    code = f'''
from sqlalchemy import create_engine, text

conn_str = {conn_str!r}
engine = create_engine(conn_str)

sql_command = {sql_command!r}
print(f"Đang thực thi: {{sql_command}}")

with engine.begin() as conn:
    result = conn.execute(text(sql_command))
    if result.returns_rows:
        rows = [dict(r._mapping) for r in result.fetchall()]
    else:
        rows = []
    row_count = len(rows) if rows else result.rowcount

if rows:
    print(f"✅ Thực thi thành công. {{len(rows)}} dòng kết quả trả về.")
else:
    if row_count == -1:
        print(f"✅ Thực thi thành công (Hệ thống DB không trả về số dòng ảnh hưởng).")
    else:
        print(f"✅ Thực thi thành công. {{row_count}} dòng bị ảnh hưởng.")

output_data = {{"result": rows, "row_count": row_count}}
'''
                    success, output, error, duration = run_python_block_sync(
                        project_id, bid, workflow_id, code, current_input,
                        timeout=7200, label=label, log_fn=log_fn, input_dir=str(input_dir),
                        stop_event=stop_event
                    )
                    if not success:
                        if log_fn:
                            log_fn("system", "error", f"❌ Thực thi SQL thất bại sau {int((datetime.now()-start).total_seconds()*1000)}ms")
                        if handle_workflow_error(error, bid, label):
                            return
                        else:
                            continue
                    else:
                        current_input = output
                        if isinstance(current_input, dict):
                            result_var = bdata.get("sqlExecResultVarName", "").strip()
                            if result_var:
                                workflow_env[result_var] = current_input.get("result")
                            row_count_var = bdata.get("sqlExecRowCountVarName", "").strip()
                            if row_count_var:
                                workflow_env[row_count_var] = current_input.get("row_count")
            elif btype == "google_sheets_read":
                url = interpolate(bdata.get("googleSheetsUrl", "")).strip()
                sheet_name = interpolate(bdata.get("googleSheetsSheetName", "")).strip()
                header_row = int(bdata.get("googleSheetsHeaderRow") or 1)
                output_var = bdata.get("outputVarName") or "sheets_data"
                custom_mappings = bdata.get("columnMappings") or {}

                if not url:
                    raise Exception("Chưa cấu hình Link Google Sheet")

                if "/spreadsheets/d/e/" in url:
                    csv_url = url.replace("/pubhtml", "/pub?output=csv").replace("/edit", "/pub?output=csv")
                    if "output=csv" not in csv_url:
                        csv_url += ("&" if "?" in csv_url else "?") + "output=csv"
                else:
                    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
                    if not match:
                        raise Exception("Link Google Sheet không hợp lệ")
                    sheet_id = match.group(1)

                    if sheet_name:
                        encoded_name = urllib.parse.quote(sheet_name)
                        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={encoded_name}"
                    else:
                        gid_match = re.search(r'[#&?]gid=([0-9]+)', url)
                        if gid_match:
                            gid = gid_match.group(1)
                            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
                        else:
                            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"

                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
                req = urllib.request.Request(csv_url, headers=headers)
                try:
                    with urllib.request.urlopen(req, timeout=20) as resp:
                        content_bytes = resp.read()
                        text = content_bytes.decode('utf-8', errors='ignore')
                except Exception as e:
                    raise Exception(f"Không thể tải Google Sheet ({e}). Hãy kiểm tra link và đảm bảo file ở chế độ Public View.")

                header_idx = max(0, header_row - 1)
                df = pd.read_csv(io.StringIO(text), header=header_idx)
                df = df.fillna('')

                clean_cols = []
                for c in df.columns:
                    c_str = str(c).strip()
                    if c_str.startswith("Unnamed:"):
                        clean_cols.append("")
                    else:
                        clean_cols.append(c_str)
                df.columns = clean_cols

                records = []
                for _, row in df.iterrows():
                    item_dict = {}
                    for orig_col in df.columns:
                        if not orig_col:
                            continue
                        val = str(row[orig_col])
                        var_key = custom_mappings.get(orig_col) or orig_col
                        item_dict[var_key] = val
                    records.append(item_dict)

                row_count_var = bdata.get("rowCountVarName") or "sheets_rows"

                if not isinstance(current_input, dict):
                    current_input = {}

                current_input[output_var] = records
                current_input[row_count_var] = len(records)
                current_input["row_count"] = len(records)

                workflow_env[output_var] = records
                workflow_env[row_count_var] = len(records)
                workflow_env["row_count"] = len(records)

                if log_fn:
                    log_fn(bid, "success", f"🟢 [Google Sheets] Đọc thành công {len(records)} dòng vào biến '{output_var}' (Số dòng: '{row_count_var}')")
            elif btype == "excel_read":
                # Đọc file Excel/CSV trong input (fallback output) → trả mảng dòng + số dòng.
                # Cùng hợp đồng output với google_sheets_read nên cắm thẳng vào khối Loop.
                file_name = interpolate(bdata.get("excelReadFile", "")).strip()
                sheet_name = interpolate(bdata.get("excelReadSheetName", "")).strip()
                header_row = int(bdata.get("excelReadHeaderRow") or 1)
                output_var = bdata.get("outputVarName") or "sheets_data"
                row_count_var = bdata.get("rowCountVarName") or "sheets_rows"
                custom_mappings = bdata.get("columnMappings") or {}

                if not file_name:
                    raise Exception("Chưa chọn/nhập tên file Excel")

                # Cho phép nhập tên biến {{...}} → sau interpolate ra tên file thật; tìm input trước, rồi output
                file_path = input_dir / file_name
                if not file_path.exists():
                    file_path = wf_dir / "output" / file_name
                if not file_path.exists():
                    raise Exception(f"Không tìm thấy file '{file_name}' trong thư mục input/output")

                header_idx = max(0, header_row - 1)
                if str(file_path).lower().endswith(".csv"):
                    df = pd.read_csv(file_path, header=header_idx)
                else:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)
                    else:
                        df = pd.read_excel(file_path, header=header_idx)
                df = df.fillna('')

                clean_cols = []
                for c in df.columns:
                    c_str = str(c).strip()
                    if c_str.startswith("Unnamed:"):
                        clean_cols.append("")
                    else:
                        clean_cols.append(c_str)
                df.columns = clean_cols

                records = []
                for _, row in df.iterrows():
                    item_dict = {}
                    for orig_col in df.columns:
                        if not orig_col:
                            continue
                        val = str(row[orig_col])
                        var_key = custom_mappings.get(orig_col) or orig_col
                        item_dict[var_key] = val
                    records.append(item_dict)

                if not isinstance(current_input, dict):
                    current_input = {}

                current_input[output_var] = records
                current_input[row_count_var] = len(records)
                current_input["row_count"] = len(records)

                workflow_env[output_var] = records
                workflow_env[row_count_var] = len(records)
                workflow_env["row_count"] = len(records)

                if log_fn:
                    log_fn(bid, "success", f"🟢 [Đọc Excel] Đọc thành công {len(records)} dòng từ '{file_name}' vào biến '{output_var}' (Số dòng: '{row_count_var}')")
            elif btype == "condition":
                logical_op = bdata.get("logicalOperator", "AND").upper()
                conditions = bdata.get("conditions")
                
                # Backward compatibility
                if not conditions and "condVariable" in bdata:
                    conditions = [{
                        "condVariable": bdata.get("condVariable"),
                        "condOperator": bdata.get("condOperator", "=="),
                        "condValue": bdata.get("condValue", "")
                    }]

                if not conditions:
                    if log_fn:
                        log_fn(bid, "warning", f"⚠️ Block [{label}] không có điều kiện nào")
                    continue_branch = False
                    continue

                if log_fn:
                    log_fn(bid, "info", f"🔀 [Condition] Kiểm tra {len(conditions)} điều kiện ({logical_op})")

                try:
                    results = []
                    for idx, cond in enumerate(conditions):
                        cond_var = cond.get("condVariable", "").strip()
                        cond_op = cond.get("condOperator", "==")
                        cond_val = cond.get("condValue", "").strip()

                        actual_val = None
                        if isinstance(current_input, dict):
                            actual_val = current_input.get(cond_var)

                        result = False
                        if cond_op == "==":
                            result = str(actual_val) == str(cond_val)
                        elif cond_op == "!=":
                            result = str(actual_val) != str(cond_val)
                        elif cond_op == ">":
                            result = float(actual_val or 0) > float(cond_val or 0)
                        elif cond_op == "<":
                            result = float(actual_val or 0) < float(cond_val or 0)
                        elif cond_op == ">=":
                            result = float(actual_val or 0) >= float(cond_val or 0)
                        elif cond_op == "<=":
                            result = float(actual_val or 0) <= float(cond_val or 0)
                        elif cond_op == "contains":
                            result = str(cond_val) in str(actual_val)
                            
                        results.append(result)
                        if log_fn:
                            log_fn(bid, "info", f"   [{idx+1}] {cond_var} ({actual_val}) {cond_op} {cond_val} ➜ {result}")
                            
                    # Calculate final result
                    if logical_op == "OR":
                        final_result = any(results)
                    else: # AND
                        final_result = all(results)
                        
                    cond_branch_taken = "true" if final_result else "false"
                    if log_fn:
                        log_fn(bid, "success", f"✅ Kết quả chung: {final_result}")
                except Exception as e:
                    if log_fn:
                        log_fn(bid, "error", f"❌ Lỗi so sánh: {e}")
                    final_status = "error"
                    break
            elif btype == "loop":
                mode = bdata.get("loopMode", "count")
                delay = float(bdata.get("loopDelay") or 0)
                
                # Check run count
                state = loop_states.setdefault(bid, {"runs": 0})

                state["runs"] += 1
                
                if not isinstance(current_input, dict):
                    current_input = {"previous_input": current_input}
                current_input["loop_iteration"] = state["runs"]
                
                if mode == "count":
                    max_count = int(bdata.get("loopCount", 0))
                    if log_fn:
                        log_fn(bid, "info", f"🔁 [Loop] Lần lặp {state['runs']} / {max_count}")
                    # "< " chứ không phải "<=": ở lần lặp thứ max_count, phải dừng NGAY (không
                    # ra lệnh chạy lại khối trước Loop thêm 1 lần thừa) - cùng lỗi off-by-one
                    # như chế độ điều kiện.
                    cond_branch_taken = "loop" if state["runs"] < max_count else "endloop"
                    if log_fn:
                        log_fn(bid, "success", f"✅ [Loop] Đi nhánh: {cond_branch_taken}")
                elif mode == "array":
                    raw_var = bdata.get("loopArrayVar", "sheets_data")
                    array_var_name = str(raw_var).strip()
                    if array_var_name.startswith("{{") and array_var_name.endswith("}}"):
                        array_var_name = array_var_name[2:-2].strip()

                    array_data = None
                    if isinstance(current_input, dict) and array_var_name in current_input:
                        array_data = current_input[array_var_name]
                    elif array_var_name in workflow_env:
                        array_data = workflow_env[array_var_name]

                    if array_data is None:
                        val = interpolate(str(raw_var))
                        if isinstance(val, list):
                            array_data = val
                        elif isinstance(val, str) and val.startswith("["):
                            try:
                                import ast
                                array_data = ast.literal_eval(val)
                            except Exception:
                                pass

                    if not isinstance(array_data, list):
                        array_data = []

                    total_items = len(array_data)
                    if log_fn:
                        log_fn(bid, "info", f"🔁 [Loop] Lần lặp {state['runs']} / {total_items} (Mảng: {array_var_name})")

                    if state["runs"] <= total_items and total_items > 0:
                        current_item = array_data[state["runs"] - 1]
                        if not isinstance(current_input, dict):
                            current_input = {}
                        
                        if isinstance(current_item, dict):
                            for k, v in current_item.items():
                                current_input[k] = v
                                workflow_env[k] = v
                        else:
                            current_input["item"] = current_item
                            workflow_env["item"] = current_item

                        cond_branch_taken = "loop"
                        if log_fn:
                            log_fn(bid, "success", f"✅ [Loop] Đi nhánh: loop (Dòng {state['runs']}/{total_items})")
                    else:
                        cond_branch_taken = "endloop"
                        if log_fn:
                            log_fn(bid, "success", f"✅ [Loop] Hoàn tất {total_items} phần tử mảng -> Đi nhánh: endloop")
                else: # condition
                    logical_op = bdata.get("logicalOperator", "AND").upper()
                    conditions = bdata.get("conditions")
                    max_count = int(bdata.get("loopMaxCount") or 0)

                    if not conditions:
                        if log_fn:
                            log_fn(bid, "warning", f"⚠️ Block [{label}] không có điều kiện nào")
                        continue_branch = False
                        continue

                    if log_fn:
                        log_fn(bid, "info", f"🔁 [Loop] Lần lặp {state['runs']}/{max_count or '∞'} - Kiểm tra {len(conditions)} điều kiện ({logical_op})")

                    try:
                        results = []
                        for idx, cond in enumerate(conditions):
                            cond_var = cond.get("condVariable", "").strip()
                            cond_op = cond.get("condOperator", "==")
                            cond_val = cond.get("condValue", "").strip()

                            actual_val = None
                            if isinstance(current_input, dict):
                                actual_val = current_input.get(cond_var)

                            result = False
                            cmp_val = cond_val
                            if isinstance(actual_val, int):
                                try: cmp_val = int(cond_val)
                                except: pass
                            elif isinstance(actual_val, float):
                                try: cmp_val = float(cond_val)
                                except: pass
                            elif isinstance(actual_val, bool):
                                cmp_val = str(cond_val).lower() == 'true'

                            if cond_op == "==":
                                result = (actual_val == cmp_val) or (str(actual_val) == str(cmp_val))
                            elif cond_op == "!=":
                                result = (actual_val != cmp_val) and (str(actual_val) != str(cmp_val))
                            elif cond_op == ">":
                                result = float(actual_val) > float(cmp_val)
                            elif cond_op == "<":
                                result = float(actual_val) < float(cmp_val)
                            elif cond_op == ">=":
                                result = float(actual_val) >= float(cmp_val)
                            elif cond_op == "<=":
                                result = float(actual_val) <= float(cmp_val)
                            elif cond_op == "contains":
                                result = str(cmp_val) in str(actual_val)

                            results.append(result)
                            if log_fn:
                                log_fn(bid, "info", f"   [{idx+1}] {cond_var} ({actual_val}) {cond_op} {cmp_val} ➜ {result}")

                        if logical_op == "OR":
                            final_result = any(results)
                        else: # AND
                            final_result = all(results)

                        if final_result:
                            # Điều kiện ĐÚNG -> true
                            cond_branch_taken = "true"
                        elif max_count and state["runs"] > max_count:
                            # loopMaxCount = số lần được PHÉP quay lại nhánh Loop (retry),
                            # KHÔNG tính lần chạy đầu tiên. VD max=2: lượt 1 và lượt 2 vẫn
                            # được phép "loop" (đúng 2 lần quay lại), chỉ lượt 3 mới bị chặn
                            # -> tổng số lần chạy khối trước Loop = max_count + 1.
                            cond_branch_taken = "endloop"
                            if log_fn:
                                log_fn(bid, "warning", f"⚠️ [Loop] Đã dùng hết {max_count} lần quay lại cho phép - dừng dù điều kiện chưa đúng")
                        else:
                            cond_branch_taken = "loop"

                        if log_fn:
                            log_fn(bid, "success", f"✅ [Loop] Kết quả chung: {final_result} -> {cond_branch_taken}")
                    except Exception as e:
                        if log_fn:
                            log_fn(bid, "error", f"❌ Lỗi so sánh vòng lặp: {e}")
                        final_status = "error"
                        break

            if continue_branch and final_status != "error":
                if btype == "loop" and cond_branch_taken == "loop":
                    delay = float(bdata.get("loopDelay") or 0)
                    if delay > 0:
                        if log_fn:
                            log_fn(bid, "info", f"⏳ [Loop] Nghỉ {delay}s trước khi lặp lại...")
                        # Chia nhỏ để check stop_event mỗi 0.5s - trước đây time.sleep(delay)
                        # nguyên khối khiến nút Dừng phải đợi hết loopDelay mới có hiệu lực.
                        _waited = 0.0
                        _step = 0.5
                        _stopped_in_loop_delay = False
                        while _waited < delay:
                            if stop_event and stop_event.is_set():
                                final_status = "stopped"
                                _stopped_in_loop_delay = True
                                break
                            _t = min(_step, delay - _waited)
                            time.sleep(_t)
                            _waited += _t
                        if _stopped_in_loop_delay:
                            break

                if isinstance(current_input, dict):
                    workflow_env.update(current_input)
                    # Nếu khối có đặt tên biến riêng (outputVarName), lưu thêm nguyên object
                    # kết quả dưới key đó - tránh bị các khối khác đè mất khi dùng chung
                    # tên key phẳng (status/result/table...) như cơ chế update() ở trên.
                    output_var_name = bdata.get("outputVarName", "").strip()
                    if output_var_name:
                        workflow_env[output_var_name] = current_input

                out_edges = edges_from.get(node_id, [])
                for e in out_edges:
                    target_id = e["target"]
                    source_handle = e.get("sourceHandle")
                    
                    # Xử lý rẽ nhánh điều kiện và vòng lặp
                    if btype in ("condition", "loop"):
                        if source_handle == cond_branch_taken:
                            queue.append((target_id, current_input))
                    else:
                        queue.append((target_id, current_input))

        total_ms = int((datetime.now() - start).total_seconds() * 1000)
        _finish_run(run_id, final_status, start)
        if log_fn:
            if final_status == "success":
                log_fn("system", "success", f"✅ Workflow hoàn thành trong {total_ms}ms")
            elif final_status == "stopped":
                log_fn("system", "warning", f"⏹ Đã dừng sau {total_ms}ms")
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        _finish_run(run_id, "error", start, error=str(e))
        if log_fn:
            log_fn("system", "error", f"❌ Lỗi hệ thống khi chạy workflow: {str(e)}")
        print(f"CRITICAL ERROR IN WORKFLOW THREAD: {err_msg}")


def _finish_run(run_id, status, start, error=None):
    finished = datetime.now()
    duration = int((finished - start).total_seconds() * 1000)
    with sqlite3.connect(str(WORKFLOW_DB)) as conn:
        conn.execute(
            "UPDATE workflow_run SET status=?, finished_at=?, duration_ms=?, error_message=? WHERE id=?",
            (status, finished.isoformat(), duration, error, run_id)
        )
    _active_runs.pop(run_id, None)
    _active_procs.pop(run_id, None)
    
    # Dọn dẹp profile duyệt web tạm thời của lượt chạy này
    profile_dir = _active_browser_profiles.pop(run_id, None)
    if profile_dir and profile_dir.parent.exists():
        import shutil
        shutil.rmtree(profile_dir.parent, ignore_errors=True)
        
    try:
        from services.browser_executor import cleanup_browser
        cleanup_browser(run_id)
    except Exception:
        pass
    # Xóa run_id khỏi workflow mapping
    for wf_id, run_set in list(_workflow_run_ids.items()):
        run_set.discard(run_id)
        if not run_set:
            _workflow_run_ids.pop(wf_id, None)

